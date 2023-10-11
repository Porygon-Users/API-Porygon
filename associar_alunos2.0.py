import openpyxl

#Função para verificar se um aluno já está alocado em alguma turma
def aluno_em_turma(planilha, aluno_chave):
    abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

    for turma in abas_turmas:
        aba_turma = planilha[turma]
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
            nome = row[0].value
            cpf = row[1].value
            aluno_turma_chave = (nome, cpf)

            if aluno_turma_chave == aluno_chave:
                return True

    return False

#Abrir a planilha
planilha = openpyxl.load_workbook('Dados Cadastrais.xlsx')

#Conjunto para manter o controle dos alunos já adicionados
alunos_adicionados = set()

while True:
    print("\nOpções:")
    print("\n1. Adicionar alunos às turmas")
    print("2. Ver alunos disponíveis e não alocados")
    print("3. Sair do programa", "\n")

    escolha = input("Escolha uma das opções: ")

    if escolha == '1':
        #Listar as abas de turma disponíveis (apenas as abertas)
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        if not abas_turmas:
            print("\nNão foram encontradas abas de turma na planilha.", "\n")
        else:
            print("\nTurmas abertas:")
            for i, turma in enumerate(abas_turmas, start=1):
                if "(fechada)" not in turma:
                    print(f"{turma.replace('(fechada)', '')}")

            #Pergunta ao usuário em qual turma deseja adicionar os alunos
            try:
                indice_turma = int(input("Digite o número da turma que deseja adicionar os alunos: ")) - 1

                if 0 <= indice_turma < len(abas_turmas) and "(fechada)" not in abas_turmas[indice_turma]:
                    turma_desejada = abas_turmas[indice_turma]
                    quantidade_alunos = int(input("Quantos alunos deseja adicionar: "))

                    #Verificar alunos na aba "Alunos" que não foram colocados a nenhuma turma
                    aba_alunos = planilha['Alunos']
                    alunos_disponiveis = []

                    for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=3):
                        nome = row[0].value
                        cpf = row[1].value
                        email = row[2].value
                        aluno_chave = (nome, cpf)

                        if not aluno_em_turma(planilha, aluno_chave):
                            alunos_disponiveis.append((nome, cpf, email))

                    #Selecionar a quantidade desejada de alunos disponíveis
                    alunos_selecionados = alunos_disponiveis[:quantidade_alunos]

                    #Adicionar os alunos selecionados à turma
                    aba_turma = planilha[turma_desejada]
                    for aluno in alunos_selecionados:
                        nome, cpf, email = aluno
                        nova_linha = [nome, cpf, email]
                        aba_turma.append(nova_linha)
                        print(f"Aluno {nome} adicionado à {turma_desejada} com sucesso.")
                        alunos_adicionados.add(aluno)

                    planilha.save('Dados Cadastrais.xlsx')
                else:
                    print("\nTurma não encontrada ou está fechada.")
            except ValueError:
                print("\nDigite um número válido para selecionar a turma.")
    elif escolha == '2':
        alunos_nao_alocados = 0

        #Verificar alunos na aba "alunos" que não foram alocados a nenhuma turma
        aba_alunos = planilha['Alunos']
        for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=3):
            nome = row[0].value
            cpf = row[1].value
            email = row[2].value
            aluno_chave = (nome, cpf)

            if not aluno_em_turma(planilha, aluno_chave):
                alunos_nao_alocados += 1

        print(f"\nAlunos disponíveis para alocação: {alunos_nao_alocados}")
    elif escolha == '3':
        break
    else:
        print("\nOpção inválida. Escolha 1 para adicionar alunos, 2 para verificar alunos disponíveis ou 3 para sair.")
print("\nAlterações realizadas com sucesso, encerrando o programa.", "\n")