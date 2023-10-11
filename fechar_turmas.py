import openpyxl

#Essa função serve para verificar se uma turma está aberta ou fechada, algo de extrema importância na hora de mostrar a lista de turmas disponíveis para o fechamento.
def turma_aberta(sheet_name):
    return "(fechada)" not in sheet_name

#Identifica o arquivo que será utilizado como banco de dados, tendo sido criado pelo código de cadastro.
arquivo_excel = "Dados Cadastrais.xlsx"

while True: #O try serve basicamente para identificar quaisquer tipo de erro que pode ocorrer, permitindo assim o encaixe de "except", que tem uma lógica parecida com  else.
    try:
        #Carrega o arquivo
        book = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError: #Aqui está o "except", que nesse caso, o erro possivel seria não encontrar o arquivo caso ele não tenha sido criado.
        print("O banco de dados não foi encontrado.") #Banco de dados -> Excel
        exit()

    #Função criada para fechar uma turma específica
    def fechar_turma(book, turma_numero): #Essa função recebe "book" que é um objeto de pasta de trabalho do Excel carregado com a biblioteca que estamos utilizando e "turma_numero" que representa o número da turma que o usuário deseja fechar.
        turma_nome = f"Turma {turma_numero}" #Ela começa criando uma string "turma_nome" que representa o nome da aba da planilha que corresponde à turma que o usuário deseja fechar, interligando a string "turma" com o número fornecido.
        if turma_nome in book.sheetnames: #A função verifica se "turma_nome" está presente na lista de nomes de abas do "book". Isso é feito para garantir que a aba da turma que deseja fechar realmente exista no arquivo excel, se o nome da aba existir, o código dentro desse IF será executado.
            sheet = book[turma_nome] #Permite o acesso e modificação da planilha referente a essa turma.
            novo_nome = f"{turma_nome} (fechada)" #Aqui atualiza o novo nome,
            sheet.title = novo_nome               #atribuindo o "fechamento".
            book.save(arquivo_excel) #Salva após alterações.
            print(f"\n{turma_nome} foi fechada com sucesso.") #Feedback
            return True #Retorna dizendo que o fechamento foi um sucesso
        else: #Se a aba não existir, retornará false e assim o código rodará novamente mostrando a lista de o input para selecionar qual turma.(próximas funções)
            print(f"\n\n\n\n\n\n\n\nDesculpe, a {turma_nome} já foi fechada ou não foi encontrada, tente novamente.")
            return False

    #Função para listar turmas disponíveis
    def listar_turmas_disponiveis(book): #Cria uma lista que itera sobre todos os "sheetnames" em "book".
        abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ') and turma_aberta(sheet)] #Atende duas condições para listar, 1º Lista apenas as "turmas" 2º Lista apenas as "turma_aberta", ou seja, sem o título de (fechada)
        print("\nTurmas disponíveis para fechamento:")
        for i, turma in enumerate(abas_turmas, start=1): #Loop para iterar sobre a lista acima, inicialmente "i" era para listar as turmas e mostrar no print abaixo, mas deixei ele listando e imprimindo somente o nome da turma correspondente.
            print(f"{turma.replace('(fechada)', '')}")
        print() #Linha em branco, como se fosse o \n

    #Pergunta ao usuário qual o número da turma que deseja fechar
    listar_turmas_disponiveis(book)
    turma_numero = int(input("Digite o número da turma que deseja fechar: "))
    
    #Verifique se a turma foi fechada com sucesso antes de continuar
    if fechar_turma(book, turma_numero):
        while True:  #Adiciona um loop interno para garantir que apenas "s" ou "n" seja aceito como resposta.
            continuar = input("\nDeseja continuar fechando turmas? (S para continuar ou N para sair): ").strip().lower()
            if continuar == 's':
                break  #Sai do loop interno e continua fechando turmas.
            elif continuar == 'n':
                book.save(arquivo_excel)
                print("\n\nAlterações realizadas com sucesso, encerrando o programa.")
                exit()  # Sai do programa.
            else:
                print("\n\nResposta inválida, por favor, tente novamente.")

print("\n") #Linha final para separar da próxima saída.