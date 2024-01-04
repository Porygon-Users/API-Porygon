import os
import pandas as pd
from openpyxl import load_workbook

def obter_caminho_para_planilha(nome_planilha):
    # Obtém o diretório atual
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))

    # Constrói o caminho completo para o arquivo Excel no diretório 'database'
    caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', f'{nome_planilha}.xlsx')

    return caminho_arquivo_excel
def listar_turmas_disponiveis(planilha_origem):
    # Obtém todas as abas da planilha de origem
    abas = planilha_origem.sheet_names

    # Filtra apenas as abas que contêm a palavra-chave "Turma"
    turmas_disponiveis = [turma for turma in abas if "Turma" in turma]

    return turmas_disponiveis
def listar_relatorios_disponiveis(planilha_destino):
    # Obtém todas as abas da planilha de destino
    abas = planilha_destino.sheet_names

    # Filtra apenas as abas que contêm a palavra-chave "Relatório"
    relatorios_disponiveis = [relatorio for relatorio in abas if "Relatório" in relatorio]

    return relatorios_disponiveis
def ler_planilha(nome_planilha):
    caminho_arquivo_excel = obter_caminho_para_planilha(nome_planilha)

    # Verifica se o arquivo Excel existe
    if os.path.exists(caminho_arquivo_excel):
        return pd.read_excel(caminho_arquivo_excel)
    else:
        # Cria um DataFrame vazio se o arquivo não existir
        return pd.DataFrame()
def salvar_planilha(nome_planilha, dados, turma_escolhida):
    caminho_arquivo_excel = obter_caminho_para_planilha(nome_planilha)

    # Salva os dados na planilha e cria uma aba com o nome "Relatório Turma X"
    with pd.ExcelWriter(caminho_arquivo_excel, engine='openpyxl', mode='a') as writer:
        # Se a aba já existe, exclua-a antes de salvar
        if turma_escolhida in writer.book.sheetnames:
            writer.book.remove(writer.book[turma_escolhida])

        # Salva os dados na nova aba
        dados.to_excel(writer, sheet_name=turma_escolhida, index=False)
        # Renomeia a aba recém-criada
        writer.book[turma_escolhida].title = f"Relatório {turma_escolhida}"
def excluir_relatorio(nome_planilha, turma_escolhida):
    caminho_arquivo_excel = obter_caminho_para_planilha(nome_planilha)

    # Exclui a aba correspondente à turma
    with pd.ExcelWriter(caminho_arquivo_excel, engine='openpyxl', mode='a') as writer:
        if turma_escolhida in writer.book.sheetnames:
            writer.book.remove(writer.book[turma_escolhida])
            print(f"Relatório da turma '{turma_escolhida}' excluído com sucesso!")
        else:
            print(f"Relatório da turma '{turma_escolhida}' não encontrado.")
def visualizar_relatorios(nome_planilha):
    caminho_arquivo_excel = obter_caminho_para_planilha(nome_planilha)
    planilha_destino = pd.ExcelFile(caminho_arquivo_excel)

    # Exibe as abas disponíveis na planilha de destino
    print("\nRelatórios existentes:")
    relatorios_disponiveis = listar_relatorios_disponiveis(planilha_destino)
    if relatorios_disponiveis:
        for relatorio in relatorios_disponiveis:
            print(f"{relatorio}")
    else:
        print("Nenhum relatório encontrado.")
def main():
    while True:
        # Exibe o menu principal
        print("\n==----------------------------==")
        print("=-----------M E N U------------=")
        print("==----------------------------==", "\n")
        print("1. Gerar Relatório")
        print("2. Excluir Relatório")
        print("3. Visualizar Relatórios")
        print("4. Sair")

        # Solicita ao usuário para escolher uma opção
        opcao = input("\nEscolha uma opção: ")

        if opcao == '1':
            # Lista todas as turmas disponíveis
            turmas_disponiveis = listar_turmas_disponiveis(planilha_origem)
            print("\nTurmas disponíveis:")
            for i, turma in enumerate(turmas_disponiveis, start=1):
                print(f"{i}. {turma}")

            # Solicita ao usuário para escolher uma turma
            escolha_turma = input("\nDigite o número da turma que deseja adicionar ao relatório (ou 'sair' para sair): ")

            if escolha_turma.lower() == 'sair':
                continue

            try:
                escolha_turma = int(escolha_turma)
                turma_escolhida = turmas_disponiveis[escolha_turma - 1]

                # Lê os dados da turma escolhida na planilha de origem
                dados_turma = planilha_origem.parse(turma_escolhida)

                # Salva os dados da turma na planilha de destino
                salvar_planilha('relatorio_de_exportacao', dados_turma, turma_escolhida)

                # Informa que a turma foi adicionada com sucesso
                print(f"\n'{turma_escolhida}' adicionada ao relatório com sucesso!")

            except (ValueError, IndexError):
                print("Escolha inválida. Por favor, digite um número de turma válido.")
                continue

        elif opcao == '2':
            # Lista todos os relatórios disponíveis para exclusão
            relatorios_disponiveis = listar_relatorios_disponiveis(planilha_destino)
            print("\nRelatórios disponíveis para exclusão:")
            for i, relatorio in enumerate(relatorios_disponiveis, start=1):
                print(f"{i}. {relatorio}")

            # Solicita ao usuário para escolher um relatório para excluir
            escolha_relatorio_excluir = input("\nDigite o número do relatório que deseja excluir (ou 'voltar' para voltar): ")

            if escolha_relatorio_excluir.lower() == 'voltar':
                continue

            try:
                escolha_relatorio_excluir = int(escolha_relatorio_excluir)
                relatorio_excluir = relatorios_disponiveis[escolha_relatorio_excluir - 1]

                # Exclui o relatório escolhido
                excluir_relatorio('relatorio_de_exportacao', relatorio_excluir)

            except (ValueError, IndexError):
                print("Escolha inválida. Por favor, digite um número de relatório válido.")
                continue

        elif opcao == '3':
            # Visualiza os relatórios existentes
            visualizar_relatorios('relatorio_de_exportacao')

        elif opcao == '4':
            print("\nSaindo do programa. Até mais!", "\n")
            break

        else:
            print("Opção inválida. Por favor, escolha uma opção válida.")
if __name__ == "__main__":
    # Lê as planilhas de origem e destino
    planilha_origem = pd.ExcelFile(obter_caminho_para_planilha('infodados'))
    planilha_destino = pd.ExcelFile(obter_caminho_para_planilha('relatorio_de_exportacao'))

    # Inicia o programa principal
    main()
