import openpyxl
import os
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from openpyxl import load_workbook

#Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__ ))

#Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

#Abrir o arquivo Excel existente ou criar um novo
if os.path.exists(caminho_arquivo_excel):
    book = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    book = openpyxl.Workbook()

#Funções
def gerenciar_alunos():    
    # Função para verificar se um aluno já está alocado em uma turma
    def aluno_em_turma_global(planilha, aluno_chave):
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        for turma in abas_turmas:
            aba_turma = planilha[turma]
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                nome = row[0].value
                cpf = row[1].value
                aluno_turma_chave = (nome, cpf)

                if aluno_turma_chave == aluno_chave:
                    return turma  # Retorna o nome da turma onde o aluno está

        return None

    # Função para listar os alunos em uma turma
    def listar_alunos_na_turma(planilha, turma_nome):
        if turma_nome in planilha.sheetnames:
            aba_turma = planilha[turma_nome]
            alunos_na_turma = []

            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                nome = row[1].value
                if nome and "(professor)" not in nome:
                    cpf = row[0].value
                    alunos_na_turma.append((cpf, nome))

            return alunos_na_turma
        else:
            return None

    # Conjunto para manter o controle dos alunos já adicionados
    alunos_adicionados = set()
    # Função para verificar se o aluno já está em alguma turma




    # Função para verificar se o aluno já está em alguma turma
    def aluno_em_turma(book, aluno_chave, turma):
        sheet = book[turma]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            aluno_na_turma = (row[0].value, row[1].value)
            if aluno_na_turma == aluno_chave:
                return True
        return False

    # Dicionário para rastrear as últimas linhas por grupo (turma)
    ultima_linha_por_grupo = {}

    # Supondo que 'book' seja a instância do Workbook
    while True:
        print("\nOpções:")
        print("\n1. Adicionar alunos às turmas")
        print("2. Ver alunos disponíveis e não alocados")
        print("3. Remover aluno de uma turma")
        print("9. Voltar", "\n")

        escolha = input("Escolha uma das opções: ")

        if escolha == '1':
            abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ')]

            if not abas_turmas:
                print("\nNão foram encontradas abas de turma na planilha.", "\n")
            else:
                print("\nTurmas abertas:")
                for i, turma in enumerate(abas_turmas, start=1):
                    if "(fechada)" not in turma:
                        print(f"{i}. {turma.replace('(fechada)', '')}")

                try:
                    indice_turma = int(input("Digite o número da turma que deseja adicionar os alunos: ")) - 1

                    if 0 <= indice_turma < len(abas_turmas) and "(fechada)" not in abas_turmas[indice_turma]:
                        turma_desejada = abas_turmas[indice_turma]
                        quantidade_alunos = int(input("Quantos alunos deseja adicionar:"))

                        aba_alunos = book['Cadastro']
                        alunos_disponiveis = []

                        for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=4):
                            nome = row[0].value
                            cpf = row[1].value
                            funcao = row[3].value
                            aluno_chave = (nome, cpf)

                            if not aluno_em_turma(book, aluno_chave, turma_desejada) and funcao == "aluno":
                                alunos_disponiveis.append((nome, cpf, row[2].value))

                        ultima_linha = ultima_linha_por_grupo.get(turma_desejada, 1)

                        # Encontrar a próxima linha vazia ou com "None" nas colunas relevantes na turma
                        aba_turma = book[turma_desejada]  # Definir aba_turma aqui
                        while True:
                            if any(aba_turma.cell(row=ultima_linha, column=col).value is None for col in [1, 2]):
                                break
                            ultima_linha += 1

                        for i, aluno in enumerate(alunos_disponiveis[:quantidade_alunos], start=1):
                            nome, cpf, email = aluno
                            proxima_linha = ultima_linha + i

                            aba_turma.cell(row=proxima_linha-1, column=1, value=nome)
                            aba_turma.cell(row=proxima_linha-1, column=2, value=cpf)

                        ultima_linha_por_grupo[turma_desejada] = proxima_linha

                        print(f"\nAlunos adicionados à {turma_desejada} com sucesso.")

                    else:
                        print("\nTurma não encontrada ou está fechada.")
                except ValueError:
                    print("\nDigite um número válido para selecionar a turma.")
            book.save(caminho_arquivo_excel)

        elif escolha == '2':
            alunos_nao_alocados = 0

            #Verificar alunos na aba "alunos" que não foram alocados a nenhuma turma
            aba_alunos = book['Cadastro']
            for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=4):
                nome = row[0].value
                cpf = row[1].value
                funcao = row[3].value
                aluno_chave = (nome, cpf)

                if not aluno_em_turma_global(book, aluno_chave) and funcao == "aluno":
                    alunos_nao_alocados += 1

            print(f"\nAlunos disponíveis para alocação: {alunos_nao_alocados}")
            book.save(caminho_arquivo_excel)

        elif escolha == '3':
            # Remover aluno de uma turma
            turma_escolhida = input("Digite o nome da turma da qual deseja remover o aluno: ")
            alunos_na_turma = listar_alunos_na_turma(book, turma_escolhida)

            if alunos_na_turma:
                print(f"\nAlunos na {turma_escolhida}:")
                for i, aluno in enumerate(alunos_na_turma, start=1):
                    nome, cpf = aluno
                    print(f"{i}. {cpf} (ID: {nome})")

                escolha_aluno = input("\nDigite o número do aluno que deseja remover: ")

                try:
                    escolha_aluno = int(escolha_aluno)
                    if 1 <= escolha_aluno <= len(alunos_na_turma):
                        aluno_remover = alunos_na_turma[escolha_aluno - 1]
                        nome_aluno, cpf_aluno = aluno_remover

                        turma_do_aluno = aluno_em_turma_global(book, (nome_aluno, cpf_aluno))
                        if turma_do_aluno:
                            aba_turma = book[turma_do_aluno]
                            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                                nome = row[0].value
                                cpf = row[1].value
                                if (nome, cpf) == (nome_aluno, cpf_aluno):
                                    aba_turma.delete_rows(row[0].row)
                                    print(f"Aluno {nome_aluno} removido da {turma_do_aluno}.")
                                    break
                        else:
                            print("Aluno não encontrado em nenhuma turma.")
                    else:
                        print("Escolha de aluno inválida.")
                except ValueError:
                    print("Escolha de aluno inválida.")
            else:
                print("Turma não encontrada ou não há alunos nessa turma.")
            book.save(caminho_arquivo_excel)

        elif escolha == '9':
            book.save(caminho_arquivo_excel)
            break
        else:
            print("\nOpção inválida. Escolha 1 para adicionar alunos, 2 para verificar alunos disponíveis, 3 para remover um aluno de uma turma, ou 9 para sair.")


    # Salve as alterações no arquivo Excel
    book.save(caminho_arquivo_excel)
def gerenciar_professores():
    def adicionar_professor_a_turma(planilha, turma_destino, professor_nome, professor_cpf):
        aba_turma = planilha[turma_destino]

        # Verificar se já existe um professor na turma
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=2):
            nome = row[0].value
            cpf = row[1].value
            if "(professor)" in cpf:
                print(f"Já existe um professor na {turma_destino}. Não é possível adicionar mais um.")
                return

        # Encontrar a primeira linha vazia após os cabeçalhos "Nome" e "CPF"
        primeira_linha_vazia = 2  # Começando na linha 2 para evitar os cabeçalhos

        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=2):
            if row[0].value is None or row[1].value is None:
                break
            primeira_linha_vazia += 1

        # Agora, você pode adicionar os dados do professor na primeira linha vazia
        aba_turma.cell(row=primeira_linha_vazia, column=1).value = professor_nome
        aba_turma.cell(row=primeira_linha_vazia, column=2).value = (f"{professor_cpf} (professor)")

        # Salvar o livro após adicionar o professor
        planilha.save(caminho_arquivo_excel)
        return True

    # Função para verificar se um professor já está alocado em uma turma
    def professor_em_turma(planilha, professor_chave):
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        for turma in abas_turmas:
            aba_turma = planilha[turma]
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
                nome = row[0].value
                cpf = row[1].value
                professor_turma_chave = (nome, cpf)

                if professor_turma_chave == professor_chave:
                    return turma  # Retorna o nome da turma onde o professor está

        return None

    def professor_em_turma_global(planilha, professor_chave):
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        for turma in abas_turmas:
            aba_turma = planilha[turma]
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                nome = row[0].value
                cpf = row[1].value
                professor_turma_chave = (nome, cpf)

                if professor_turma_chave == professor_chave:
                    return turma  # Retorna o nome da turma onde o aluno está

        return None

    # Função para listar os professores em uma turma
    def listar_professores_na_turma(planilha, turma_nome):
        if turma_nome in planilha.sheetnames:
            aba_turma = planilha[turma_nome]
            professores_na_turma = []
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                nome = row[1].value
                if nome and "(professor)" in nome:
                    cpf = row[0].value
                    professores_na_turma.append((cpf, nome))
            return professores_na_turma
        else:
            return None

    # Conjunto para manter o controle dos professores já adicionados
    professores_adicionados = set()

    while True:
        print("\nOpções:")
        print("\n1. Adicionar professores às turmas")
        print("2. Ver professores disponíveis e não alocados")
        print("3. Remover professor de uma turma")
        print("9. Voltar", "\n")

        escolha = input("Escolha uma das opções: ")

        if escolha == '1':
            # Listar as abas de turma disponíveis (apenas as abertas)
            abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ')]

            if not abas_turmas:
                print("\nNão foram encontradas abas de turma na planilha.", "\n")
            else:
                print("\nTurmas abertas:")
                for i, turma in enumerate(abas_turmas, start=1):
                    if "(fechada)" not in turma:
                        print(f"{i}. {turma.replace('(fechada)', '')}")

                # Pergunta ao usuário em qual turma deseja adicionar os professores
                try:
                    indice_turma = int(input("\nDigite o número da turma que deseja adicionar os professores: ")) - 1

                    if 0 <= indice_turma < len(abas_turmas) and "(fechada)" not in abas_turmas[indice_turma]:
                        turma_desejada = abas_turmas[indice_turma]

                        # Verificar professores na aba "Cadastro" que não foram colocados em nenhuma turma
                        aba_professores = book['Cadastro']
                        professores_disponiveis = []

                        for row in aba_professores.iter_rows(min_row=2, max_row=aba_professores.max_row, min_col=1, max_col=4):
                            nome = row[0].value
                            cpf = row[1].value
                            funcao = row[3].value
                            professor_chave = (nome, cpf)

                            if not professor_em_turma(book, professor_chave) and funcao == "professor":
                                professores_disponiveis.append((nome, cpf))

                        # Listar os professores disponíveis
                        print("\nProfessores disponíveis:")
                        for i, professor in enumerate(professores_disponiveis, start=1):
                            nome, cpf = professor
                            print(f"{i}. {cpf} (ID: {nome})")

                        # Perguntar quais professores adicionar
                        escolha_professores = input("\nDigite o número do professor que deseja adicionar: ")
                        indices_professores = [int(index.strip()) for index in escolha_professores.split(',')]

                        # Adicionar os professores selecionados à turma
                        aba_turma = book[turma_desejada]
                        for indice_professor in indices_professores:
                            if 1 <= indice_professor <= len(professores_disponiveis):
                                professor = professores_disponiveis[indice_professor - 1]
                                nome, cpf = professor
                                if adicionar_professor_a_turma(book, turma_desejada, nome, cpf):
                                    print(f"Professor {cpf} adicionado à {turma_desejada} com sucesso.")
                                else:
                                    pass
                                professores_adicionados.add(professor)
                            else:
                                print(f"Índice {indice_professor} inválido.")

                    else:
                        print("\nTurma não encontrada ou está fechada.")
                except ValueError:
                    print("\nDigite um número válido para selecionar a turma.")
            book.save(caminho_arquivo_excel)

        elif escolha == '2':
            professores_nao_alocados = 0

            # Verificar professores na aba "Cadastro" que não foram alocados a nenhuma turma
            aba_professores = book['Cadastro']
            for row in aba_professores.iter_rows(min_row=2, max_row=aba_professores.max_row, min_col=1, max_col=4):
                nome = row[0].value
                cpf = row[1].value
                funcao = row[3].value
                professor_chave = (nome, cpf)

                if not professor_em_turma(book, professor_chave) and funcao == "professor":
                    professores_nao_alocados += 1

            print(f"\nProfessores disponíveis para alocação: {professores_nao_alocados}")
            book.save(caminho_arquivo_excel)

        elif escolha == '3':
            # Remover professor de uma turma
            turma_escolhida = input("Digite o nome da turma da qual deseja remover o professor: ")
            professores_na_turma = listar_professores_na_turma(book, turma_escolhida)

            if professores_na_turma:
                print(f"\nProfessor na {turma_escolhida}:")
                for i, professor in enumerate(professores_na_turma, start=1):
                    nome, cpf = professor
                    print(f"{i}. {cpf}")

                escolha_professor = input("\nDigite o número do professor que deseja remover: ")

                try:
                    escolha_professor = int(escolha_professor)
                    if 1 <= escolha_professor <= len(professores_na_turma):
                        professor_remover = professores_na_turma[escolha_professor - 1]
                        nome_professor, cpf_professor = professor_remover

                        turma_do_professor = professor_em_turma_global(book, (nome_professor, cpf_professor))
                        if turma_do_professor:
                            aba_turma = book[turma_do_professor]
                            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                                nome = row[0].value
                                cpf = row[1].value
                                if (nome, cpf) == (nome_professor, cpf_professor):
                                    aba_turma.delete_rows(row[0].row)
                                    print(f"{cpf_professor} removido da {turma_do_professor}.")
                                    break
                        else:
                            print("Professor não encontrado em nenhuma turma.")
                except ValueError:
                    print("Escolha de professor inválida.")
                book.save(caminho_arquivo_excel)

        elif escolha == '9':
            book.save(caminho_arquivo_excel)
            break
        else:
            print("\nOpção inválida. Escolha 1 para adicionar professores, 2 para verificar professores disponíveis, 3 para remover um professor de uma turma, ou 9 para sair.")


    # Salve as alterações no arquivo Excel
    book.save(caminho_arquivo_excel)
def grupos():
    # Função para criar grupos em uma turma
    def criar_grupos(planilha, turma_nome, num_alunos_por_grupo):
        # Carregar o arquivo da planilha
        wb = openpyxl.load_workbook(planilha)
        
        # Selecionar a aba da turma
        try:
            sheet = wb[turma_nome]
        except KeyError:
            print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
            return
        
        # Obter o índice da coluna que contém o cabeçalho "Grupos"
        coluna_grupos = None
        for cell in sheet[1]:  # Percorre as células da primeira linha
            if cell.value == "Grupos":
                coluna_grupos = cell.column_letter
                break
        
        if coluna_grupos is None:
            print("Não foi encontrada uma coluna com o cabeçalho 'Grupos'.")
            return
        
        # Obter a lista de alunos na turma
        alunos = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row if cell.value]
        
        # Exibir a quantidade de alunos na turma
        num_alunos = len(alunos)
        print(f"\nA {turma_nome} possui {num_alunos} alunos.")
        
        # Calcular o número total de grupos
        num_grupos = num_alunos // num_alunos_por_grupo
        alunos_restantes = num_alunos % num_alunos_por_grupo
        
        # Criar grupos com o número especificado de alunos
        grupos = []
        for i in range(num_grupos):
            grupo_alunos = alunos[i * num_alunos_por_grupo: (i + 1) * num_alunos_por_grupo]
            grupos.append(f'Grupo {i + 1} - {", ".join(grupo_alunos)}')
        
        # Se houver alunos restantes, criar o último grupo
        if alunos_restantes > 0:
            grupo_alunos = alunos[-alunos_restantes:]
            grupos.append(f'Grupo {num_grupos + 1} - {", ".join(grupo_alunos)}')
        
        # Adicionar os grupos à coluna "Grupos" na planilha
        for i, grupo in enumerate(grupos):
            sheet[f"{coluna_grupos}{i + 2}"] = grupo
        
        # Salvar a planilha
        wb.save(planilha)
        print(f"\nForam criados {len(grupos)} grupos com {num_alunos_por_grupo} alunos cada{' e um grupo final com ' + str(alunos_restantes) + ' alunos' if alunos_restantes > 0 else ''} na '{turma_nome}'.")

    # Função para listar as turmas existentes
    def listar_turmas(planilha):
        wb = openpyxl.load_workbook(planilha)
        turmas = [sheet for sheet in wb.sheetnames if sheet.startswith('Turma ')]
        if turmas:
            print("\nTurmas existentes:", "\n")
            for turma in turmas:
                print(turma)
        else:
            print("\nNão foram encontradas abas de turma na planilha.")

    # Função para contar alunos em uma turma
    def contar_alunos(planilha, turma_nome):
        # Carregar o arquivo da planilha
        wb = openpyxl.load_workbook(planilha)
        
        # Selecionar a aba da turma
        try:
            sheet = wb[turma_nome]
        except KeyError:
            print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
            return
        
        # Contar alunos na turma
        num_alunos = sum(1 for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) if row[0].value)
        return num_alunos  # Retornar a quantidade de alunos na turma

    # Função principal do programa
    def main():
        while True:
            print("\nOpções:")
            print("\n1. Criar grupos")
            print("2. Listar turmas existentes")
            print("9. Voltar", "\n")
            
            escolha = input("Escolha uma das opções: ")
            
            if escolha == '1':
                turma = input('\nDigite o nome da turma: ')
                alunos_na_turma = contar_alunos(caminho_arquivo_excel, turma)  # Mostrar a quantidade de alunos na turma
                if alunos_na_turma is not None:
                    if alunos_na_turma > 0:
                        print(f"\nA turma '{turma}' possui {alunos_na_turma} alunos.")
                        alunos_por_grupo = int(input("\nDigite o número de alunos por grupo: "))
                        criar_grupos(caminho_arquivo_excel, turma, alunos_por_grupo)
                book.save(caminho_arquivo_excel)
            elif escolha == '2':
                listar_turmas(caminho_arquivo_excel)
                book.save(caminho_arquivo_excel)
            elif escolha == '9':
                book.save(caminho_arquivo_excel)
                break
            else:
                print("\nOpção inválida. Tente novamente.", "\n")

    if __name__ == "__main__":
        main()
def turmas():
    def criar_nova_turma(book, nome_turma):
        nome_aba = f"Turma {nome_turma}"
        nova_aba = book.create_sheet(nome_aba)
        cabecalhos = ["ID", "NOME", "Grupos", "Inicio do Curso", "Fim do Curso"]
        nova_aba.append(cabecalhos)

        for col_idx, header in enumerate(cabecalhos, 1):
            coluna_letra = get_column_letter(col_idx)
            cell = nova_aba[f"{coluna_letra}1"]
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
            cell.value = header
            nova_aba.column_dimensions[coluna_letra].width = 20

        print(f"\nA turma '{nome_turma}' foi criada com sucesso.")
        return nome_aba



    def adicionar_data_e_ciclos(planilha, turma_destino):
        aba_turma = planilha[turma_destino]
        ciclos = []

        while True:
            try:
                data_inicio = datetime.strptime(input("\nDigite a data de início do curso (DD/MM/AAAA): "), "%d/%m/%Y")
                data_fim = datetime.strptime(input("Digite a data de término do curso (DD/MM/AAAA): "), "%d/%m/%Y")
                if data_fim < data_inicio:
                    print("A data de término é anterior à data de início, tente novamente")
                else:
                    break
            except ValueError:
                print("Formato de data inválido. Use o formato DD/MM/AAAA")

        aba_turma.cell(row=1, column=4).value = "Início do Curso"
        aba_turma.cell(row=1, column=5).value = "Fim do Curso"
        aba_turma.cell(row=2, column=4).value = data_inicio.strftime('%d/%m/%Y')
        aba_turma.cell(row=2, column=5).value = data_fim.strftime('%d/%m/%Y')

        qtd_ciclos = int(input("\nQuantos ciclos você deseja: "))
        for i in range(qtd_ciclos):
            peso_ciclos = float(input(f"Digite o peso para o ciclo {i + 1}: "))
            coluna_peso = 5 + qtd_ciclos + i  # Calcula a coluna para o peso do ciclo atual
            aba_turma.cell(row=1, column=coluna_peso+1).value = f"PESO C{i + 1}"
            aba_turma.cell(row=2, column=coluna_peso+1).value = peso_ciclos

        while True:
            choice_cycle_type = input("\nEscolha o tipo do ciclo:\n\n1-Simétrico\n2-Definir cada ciclo\n\nEscolha uma das opções: ")
            if choice_cycle_type == "1" or choice_cycle_type == "2":
                break
            else:
                print("Opção inválida, tente novamente")

        if choice_cycle_type == "1":
            duracao_ciclo = (data_fim - data_inicio) / qtd_ciclos

            for i in range(qtd_ciclos):
                ciclo_nome = f"Ciclo {i + 1} (Início/Fim)"
                ciclo_inicio = data_inicio + i * duracao_ciclo
                ciclo_fim = ciclo_inicio + duracao_ciclo - timedelta(days=1)
                ciclos.append((ciclo_nome, ciclo_inicio, ciclo_fim))

            # Adiciona os nomes dos ciclos e as datas de início e término abaixo de cada ciclo
            for i, ciclo in enumerate(ciclos, start=1):
                ciclo_nome, ciclo_inicio, ciclo_fim = ciclo
                coluna_ciclo = 5 + (i)  # Calcula a coluna para o ciclo atual
                aba_turma.cell(row=1, column=coluna_ciclo).value = f"{ciclo_nome}"
                aba_turma.cell(row=2, column=coluna_ciclo).value = ciclo_inicio.strftime('%d/%m/%Y')
                aba_turma.cell(row=3, column=coluna_ciclo).value = ciclo_fim.strftime('%d/%m/%Y')

        
        elif choice_cycle_type == "2":
            while True:
                try:
                    duracao_ciclo = (data_fim - data_inicio)
                    ciclo_datas = []  # Para armazenar as datas de início dos ciclos
                    ciclo_datas_fim = []  # Para armazenar as datas de término dos ciclos

                    for i in range(qtd_ciclos):
                        ciclo_nome = f"Ciclo {i + 1}"
                        while True:
                            try:
                                ciclo_inicio = datetime.strptime(input(f"\nDigite a data de início do {ciclo_nome} (DD/MM/AAAA): "), "%d/%m/%Y")
                                if ciclo_inicio < data_inicio or ciclo_inicio > data_fim:
                                    print("Data está fora do ciclo de curso, tente novamente")
                                elif ciclo_inicio in ciclo_datas:
                                    print("A data de início do ciclo já foi escolhida antes, tente novamente")
                                elif ciclo_inicio < max(ciclo_datas_fim, default=data_inicio):
                                    print("A data de início do ciclo deve ser posterior à data de término do ciclo anterior.")
                                else:
                                    break
                            except ValueError:
                                print("Formato de data inválido. Use o formato DD/MM/AAAA.")
    
                        while True:
                            try:
                                ciclo_fim = datetime.strptime(input(f"Digite a data de finalização do {ciclo_nome} (DD/MM/AAAA): "), "%d/%m/%Y")
                                if ciclo_fim < data_inicio or ciclo_fim > data_fim:
                                    print("Data está fora do ciclo de curso, tente novamente")
                                elif ciclo_fim in ciclo_datas or ciclo_fim in ciclo_datas_fim:
                                    print("A data de término do ciclo já foi escolhida antes, tente novamente")
                                elif ciclo_fim < ciclo_inicio:
                                    print("A data de término do ciclo deve ser posterior à data de início do ciclo.")
                                else:
                                    break
                            except ValueError:
                                print("Formato de data inválido. Use o formato DD/MM/AAAA.")

                        ciclos.append((ciclo_nome, ciclo_inicio, ciclo_fim))
                        ciclo_datas.append(ciclo_inicio)
                        ciclo_datas_fim.append(ciclo_fim)

                    break
                except ValueError:
                    print("Formato de data inválido. Use o formato DD/MM/AAAA.")

        # Salvar a planilha no arquivo infodados.xlsx
        caminho_arquivo_excel = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'database', 'infodados.xlsx')
        planilha.save(caminho_arquivo_excel)
        print("\n==--Turma criada com sucesso!--==")

    # Defina as funções ausentes
    def mostrar_numero_de_turmas(book):
        turmas_existentes = [sheet for sheet in book.sheetnames if sheet.startswith("Turma ")]
        if not turmas_existentes:
            print("\nNenhuma turma encontrada.")
        else:
            print("\nTurmas existentes:")
            for turma in turmas_existentes:
                print("\n", turma)
            print(f"\nTotal de turmas: {len(turmas_existentes)}")

    def excluir_turmas(book, turmas_a_excluir):
        for turma_nome in turmas_a_excluir:
            if turma_nome in book.sheetnames:
                book.remove(book[turma_nome])
                print(f"\n{turma_nome} excluída com sucesso.")
            else:
                print(f"\nA turma {turma_nome} não foi encontrada.")
        book.save(caminho_arquivo_excel)

    def listar_turmas_disponiveis(book):
        abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ') and not sheet.endswith('(fechada)')]
        print("\nTurmas disponíveis para fechamento:")
        for i, turma in enumerate(abas_turmas, start=1):
            print(f"{turma.replace('(fechada)', '')}")
        print()

    def fechar_turma(book, nome_turma):
        turma_nome = nome_turma
        if turma_nome in book.sheetnames:
            sheet = book[turma_nome]
            novo_nome = f"{turma_nome} (fechada)"
            sheet.title = novo_nome
            book.save(caminho_arquivo_excel)
            print(f"\n{turma_nome} foi fechada com sucesso.")
            return True
        else:
            print(f"\n\n\n\nDesculpe, a {turma_nome} já foi fechada ou não foi encontrada, tente novamente.")
            return False

    def menu():
        print("\n1 - Criar nova turma")
        print("2 - Visualizar turmas")
        print("3 - Excluir turmas")
        print("4 - Fechar turma")
        print("9 - Voltar", "\n")

    while True:
        menu()
        opcao = input('Digite o número da opção: ')

        if opcao == "1":
            nome_turma = input("Digite um nome para a nova turma: ")
            turma_destino = criar_nova_turma(book, nome_turma)
            adicionar_data_e_ciclos(book, turma_destino)
        elif opcao == "2":
            mostrar_numero_de_turmas(book)
        elif opcao == "3":
            listar_turmas_disponiveis(book)
            turmas_a_excluir = input("Digite o nome das turmas que deseja excluir (separadas por vírgula): ")
            turmas_a_excluir = [turma.strip() for turma in turmas_a_excluir.split(",")]
            excluir_turmas(book, turmas_a_excluir)
        elif opcao == "4":
            listar_turmas_disponiveis(book)
            nome_turma = input("Digite o nome da turma que deseja fechar: ")
            fechar_turma(book, nome_turma)
        elif opcao == "9":
            break
        else:
            print("\nOpção inválida.", "\n")


    # Salve as alterações no arquivo Excel
    book.save(caminho_arquivo_excel)
def gerenciar_notas():
    def listar_alunos_na_turma(planilha, turma_nome):
        if turma_nome in planilha.sheetnames:
            aba_turma = planilha[turma_nome]
            alunos_na_turma = []

            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                nome = row[1].value
                if nome and "(professor)" not in nome:
                    cpf = row[0].value
                    alunos_na_turma.append((nome, cpf))

            return alunos_na_turma
        else:
            return None

    def mostrar_numero_de_turmas(book):
        turmas_existentes = [sheet for sheet in book.sheetnames if sheet.startswith("Turma ")]
        if not turmas_existentes:
            print("\nNenhuma turma encontrada.")
        else:
            print("\nTurmas existentes:")
            for i, turma in enumerate(turmas_existentes, start=1):
                print(f"{i}. {turma}")
            print(f"\nTotal de turmas: {len(turmas_existentes)}")
            return turmas_existentes

    # Para rastrear as linhas dos alunos
    dict_linhas_alunos = {}

    # Listar turmas disponíveis
    turmas_disponiveis = mostrar_numero_de_turmas(book)

    # Solicitar o número da turma
    escolha_numero_turma = input("Digite o número da turma para gerenciar scores: ")

    try:
        escolha_numero_turma = int(escolha_numero_turma)
        if 1 <= escolha_numero_turma <= len(turmas_disponiveis):
            nome_turma_notas = turmas_disponiveis[escolha_numero_turma - 1]

            # Obter a aba da turma
            aba_turma_notas = book[nome_turma_notas]

            # Listar alunos na turma
            alunos_na_turma = listar_alunos_na_turma(book, nome_turma_notas)
            if alunos_na_turma:
                print("\nAlunos na turma:")
                for i, aluno in enumerate(alunos_na_turma, start=1):
                    nome_aluno, cpf_aluno = aluno
                    print(f"{i}. {nome_aluno} (ID: {cpf_aluno})")

                    # Adiciona a linha do aluno no dicionário
                    dict_linhas_alunos[cpf_aluno] = i + 1  # +1 para considerar o cabeçalho

                # Solicitar o número do aluno para gerenciar notas
                escolha_aluno_notas = input("\nDigite o número do aluno para gerenciar scores: ")

                try:
                    escolha_aluno_notas = int(escolha_aluno_notas)
                    if 1 <= escolha_aluno_notas <= len(alunos_na_turma):
                        aluno_notas = alunos_na_turma[escolha_aluno_notas - 1]
                        nome_aluno, cpf_aluno = aluno_notas

                        # Encontrar colunas de ciclos
                        colunas_ciclos = [coluna for coluna in aba_turma_notas[1] if coluna.value and "Ciclo" in coluna.value]

                        if colunas_ciclos:
                            qtd_ciclos = len(colunas_ciclos)
                            for i, coluna_ciclo in enumerate(colunas_ciclos):
                                ciclo_nome = coluna_ciclo.value.split()[1]
                                peso_ciclo = aba_turma_notas.cell(row=2, column=coluna_ciclo.col_idx).value
                                nota = float(input(f"Digite o score para o Ciclo {ciclo_nome} do aluno {nome_aluno}: "))
                                # Obtém a linha correta do aluno
                                linha_aluno = dict_linhas_alunos[cpf_aluno]

                                # Calcular a coluna correta para as notas
                                coluna_nota = 5 + qtd_ciclos*2 + i  # Após as colunas de peso e ciclos

                                # Adicionar título da nota e valor na célula correspondente
                                aba_turma_notas.cell(row=1, column=coluna_nota+1).value = f"SCORE C{ciclo_nome}"
                                aba_turma_notas.cell(row=linha_aluno, column=coluna_nota+1).value = nota

                            print(f"\nScore do aluno '{cpf_aluno}' adicionadas/alteradas com sucesso.")
                        else:
                            print("\nA turma não possui ciclos definidos.")
                    else:
                        print("\nEscolha de aluno inválida.")
                except ValueError:
                    print("\nEscolha de aluno inválida.")
            else:
                print("\nNão há alunos na turma ou a turma não existe.")
        else:
            print("\nEscolha de turma inválida.")
    except ValueError:
        print("\nEscolha de turma inválida.")
    #Encontrar a última coluna de notas
    ultima_coluna_nota = coluna_nota + 1  # A última coluna onde uma nota foi adicionada

    #Adicionar título da coluna "MÉDIAS" na célula correspondente
    aba_turma_notas.cell(row=1, column=ultima_coluna_nota + 1).value = "FEE"

    #Calcular as médias ponderadas e adicionar na coluna "MÉDIAS"
    for aluno, linha in dict_linhas_alunos.items():
        soma_notas_ponderadas = 0
        soma_pesos = 0

        # Calcular a média ponderada para cada ciclo
        for i, coluna_ciclo in enumerate(colunas_ciclos):
            peso_ciclo = aba_turma_notas.cell(row=2, column=coluna_ciclo.col_idx + qtd_ciclos).value
            nota_ciclo = aba_turma_notas.cell(row=linha, column=coluna_ciclo.col_idx + qtd_ciclos * 2).value

            # Certifique-se de que peso_ciclo seja tratado como um número
            peso_ciclo = float(peso_ciclo) if peso_ciclo is not None else 0.0

            if nota_ciclo is not None:
                soma_notas_ponderadas += peso_ciclo * float(nota_ciclo)
                soma_pesos += peso_ciclo

        # Calcular a média ponderada total
        if soma_pesos != 0:
            media_ponderada1 = soma_notas_ponderadas / soma_pesos
            media_ponderada = round(media_ponderada1, 2)
            aba_turma_notas.cell(row=linha, column=ultima_coluna_nota + 1).value = media_ponderada
    book.save(caminho_arquivo_excel)
def listar_medias(book):
    while True:
        try:
            #Obter todas as turmas disponíveis
            turmas = [turma for turma in book.sheetnames if "Turma" in turma]

            #Imprimir as turmas disponíveis e pedir ao usuário para entrar com o número da turma
            print("\nTurmas disponíveis:")
            for i, turma in enumerate(turmas, start=1):
                print(f"{i}. {turma}")
            turma_selecionada = turmas[int(input("\nEntre com o número da turma: ")) - 1]

            #Selecionar a aba correspondente à turma selecionada pelo usuário
            sheet = book[turma_selecionada]

            #Inicializar a lista de médias
            lista_medias = []

            #Inicializar coluna_medias com um valor padrão
            coluna_medias = None

            #Identificar a coluna "MÉDIAS"
            for cell in sheet[1]:
                if cell.value == "FEE":
                    coluna_medias = cell.column
                    break

            #Verificar se a coluna "MÉDIAS" foi encontrada
            if coluna_medias is None:
                print("\Score não localizado, por favor, realize uma primeira atribuição antes de listar o FEE (Fator de Ensino Evolutivo).")
                return []

            #Percorrer cada linha da aba selecionada
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print()
                #Verificar se a célula da coluna "MÉDIAS" está vazia
                if row[coluna_medias - 1] is None:
                    #Adicionar à lista_medias uma string formatada com o ID do aluno, nome e "Média ainda não calculada"
                    lista_medias.append(f"ID: {row[0]} - {row[1]}: Score não atribuido")
                else:
                    #Adicionar à lista_medias uma string formatada com o ID do aluno, nome e média
                    lista_medias.append(f"ID: {row[0]} - {row[1]}: {row[coluna_medias - 1]}")

            # Retornar lista_medias
            return lista_medias
            break
        except (ValueError, IndexError, TypeError):
            print("\nPor favor, entre com um número de turma válido.")
        
#Menu
while True:
    try:
        print("\n==----------------------------==")
        print("=-----------M E N U------------=")
        print("==----------------------------==")
        print("\n1. Gerenciar alunos e professores")
        print("2. Gerenciar turmas")
        print("3. Gerenciar grupos")
        print("4. Gerenciar scores")
        print("5. Sair do programa", "\n")
        escolha = int(input("Escolha uma das opções: "))

        while True:
            if escolha == 1:
                print("\n\nOpções:")
                print("\n1. Gerenciar alunos")
                print("2. Gerenciar professores")
                print("9. Voltar", "\n")
                escolha1 = input("Escolha uma das opções: ")
                if escolha1 == '1':
                    gerenciar_alunos()
                elif escolha1 == '2':
                    gerenciar_professores()
                elif escolha1 == '9':
                    book.save(caminho_arquivo_excel)
                    break
                else:
                    print("\nOpção inválida, tente novamente.")
            elif escolha == 2:
                    turmas()
                    book.save(caminho_arquivo_excel)
                    break
            elif escolha == 3:
                    book.save(caminho_arquivo_excel)
                    grupos()
                    break
            elif escolha == 4:
                print("\n\nOpções:")
                print("\n1. Atribuir Scores")
                print("2. Visualizar FEE (Fator de Ensino Evolutivo)")
                print("9. Voltar", "\n")
                escolha4 = input("Escolha uma das opções: ")
                if escolha4 == '1':
                    gerenciar_notas()
                elif escolha4 == '2':
                    print('\n'.join(listar_medias(book)))
                elif escolha4 == '9':
                    book.save(caminho_arquivo_excel)
                    break
                
            elif escolha == 5:
                print("\n\nAlterações realizadas com sucesso, encerrando o programa.", "\n")
                book.save(caminho_arquivo_excel)
                exit()
            else:
                print("\n\n\n\nOpção inválida, tente novamente.")
                break
    except ValueError:
        print("\n\nPor favor, entre apenas com números.")

#Salva as alterações no arquivo Excel
book.save(caminho_arquivo_excel)