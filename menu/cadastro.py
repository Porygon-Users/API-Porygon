import openpyxl
import os
import random

def gerar_id():
    return ''.join(random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890') for _ in range(3))

def gerar_id_aluno(ids_usados):
    while True:
        id = 'A' + gerar_id()
        if id not in ids_usados:
            ids_usados.add(id)
            return id

def gerar_id_professor(ids_usados):
    while True:
        id = 'P' + gerar_id()
        if id not in ids_usados:
            ids_usados.add(id)
            return id

def validar_cpf(cpf):
    if cpf.isdigit() and len(cpf) == 11:
        return True
    else:
        return False

def formatar_cpf(cpf):
    return f"{cpf[:9]}-{cpf[9:]}"

def cadastrar_usuario(planilha, funcao, ids_usados):
    nome = input("Digite o nome: ").lower()

    if funcao == 'aluno':
        id = gerar_id_aluno(ids_usados)
    else:
        id = gerar_id_professor(ids_usados)

    cpf_valido = False  
    while not cpf_valido:  
        cpf = input("Digite o CPF: ")
        if validar_cpf(cpf):
            cpf = formatar_cpf(cpf)  
            cpf_valido = True  
        else:
            print("\nCPF inválido, tente novamente.")

    proxima_linha = planilha.max_row + 1

    planilha.cell(row=proxima_linha, column=1, value=id)
    planilha.cell(row=proxima_linha, column=4, value=funcao)
    planilha.cell(row=proxima_linha, column=2, value=nome)
    planilha.cell(row=proxima_linha, column=3, value=cpf)

    print("\nBem vindo(a) a PBLTeX.", "\n")
    print("Cadastro feito com sucesso!", "\n")

diretorio_atual = os.path.dirname(os.path.abspath(__file__))

caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

if os.path.exists(caminho_arquivo_excel):
    arquivo_excel = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    arquivo_excel = openpyxl.Workbook()

if "Cadastro" in arquivo_excel.sheetnames:
    planilha = arquivo_excel["Cadastro"]
else:
    planilha = arquivo_excel.create_sheet("Cadastro")
    planilha.append(["ID", "Nome", "CPF", "Função"])

ids_usados = set()

while True:
    print("\nBem vindo(a) ao sistema de cadastro da PBLTeX!", "\n")
    print("Escolha a opção:")
    print("1. Cadastrar como Aluno")
    print("2. Cadastrar como Professor")
    print("9. Sair do programa")

    opcao = input("Opção: ")

    if opcao == '9':
        print("\nCadastro Feito com Sucesso, Até logo!", "\n")
        break
    elif opcao in ['1', '2']:
        funcao = 'aluno' if opcao == '1' else 'professor'
        cadastrar_usuario(planilha, funcao, ids_usados)

        adicionar_outro = input("Gostaria de adicionar outro aluno/professor? (s/n): ").lower()
        if adicionar_outro != 's':
            print("\nDados atualizados com sucesso, encerrando o programa. Até logo!", "\n")
            break  
    else:
        print("Opção inválida. Tente novamente.")
        continue

arquivo_excel.save(caminho_arquivo_excel)
arquivo_excel.close()
