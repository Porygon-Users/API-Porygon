import os
import openpyxl

def criar_aba_notas(book):
    if 'Notas' not in book.sheetnames:
        book.create_sheet('Notas')
        notas_sheet = book['Notas']
        notas_sheet.append(["Aluno"])
    if 'Pesos' not in book.sheetnames:
        book.create_sheet('Pesos')
        pesos_sheet = book['Pesos']
        pesos_sheet.append(["Ciclo", "Peso"])

def listar_turmas(book):
    turmas = [sheet.title for sheet in book.worksheets if sheet.title.startswith('Turma')]
    print("Turmas disponíveis:")
    for i, turma in enumerate(turmas, 1):
        print(f"{turma}")
    return turmas

def listar_alunos_em_turma(turma_sheet):
    alunos = [row[0].value for row in turma_sheet.iter_rows(min_row=2, max_row=turma_sheet.max_row, min_col=1, max_col=1)]
    return alunos

def transferir_aluno_para_notas(book, turma, aluno, ciclo_atual, nota):
    notas_sheet = book['Notas']
    ciclo_atual_str = f'Ciclo {ciclo_atual}'

    if ciclo_atual_str not in [cell.value for cell in notas_sheet[1]]:
        notas_sheet.cell(row=1, column=notas_sheet.max_column + 1, value=ciclo_atual_str)

    ciclo_col = [cell.value for cell in notas_sheet[1]].index(ciclo_atual_str) + 1

    aluno_row = None
    for row in notas_sheet.iter_rows(min_row=2, max_row=notas_sheet.max_row, min_col=1, max_col=1):
        if row[0].value == aluno:
            aluno_row = row[0].row
            break

    if aluno_row is not None:
        # Se o aluno já existe na planilha, atualize a nota na coluna do ciclo atual
        notas_sheet.cell(row=aluno_row, column=ciclo_col, value=nota)
    else:
        # Se o aluno não existe na planilha, adicione uma nova linha para o aluno e atribua a nota
        notas_sheet.append([aluno] + [None] * (ciclo_col - 2) + [nota])

def atribuir_pesos(book):
    notas_sheet = book['Notas']
    ciclos_na_notas = [cell.value for cell in notas_sheet[1] if cell.value.startswith('Ciclo ')]
    pesos_sheet = book['Pesos']

    # Criar um dicionário para mapear ciclos sem peso para índices originais
    ciclos_sem_peso_dict = {ciclo: i for i, ciclo in enumerate(ciclos_na_notas) if not any(ciclo in cell.value for cell in pesos_sheet['A'] if cell.value is not None)}

    # Verificar se existem ciclos sem peso
    if not ciclos_sem_peso_dict:
        print("\n\nNão foram encontrados ciclos sem peso, tente atribuir novas notas.")
        return  # Retorna ao menu principal

    # Mostra os ciclos sem peso
    print("\n\nCiclos sem peso:")
    for i, ciclo in enumerate(ciclos_na_notas, 1):
        if ciclo in ciclos_sem_peso_dict:
            print(f"{i}. {ciclo}")

    while True:
        ciclo_idx = int(input("\nDigite o número do ciclo para atribuir peso: ")) - 1

        if 0 <= ciclo_idx < len(ciclos_na_notas):
            ciclo_selecionado = ciclos_na_notas[ciclo_idx]

            # Verificar se o ciclo já possui peso
            if ciclo_selecionado not in ciclos_sem_peso_dict:
                print("Ciclo inexistente ou já possui peso. Tente novamente.")
                continue

            peso = float(input(f"Digite o peso para {ciclo_selecionado}: "))

            pesos_sheet.append([ciclo_selecionado, peso])

            # Remover o ciclo do dicionário
            del ciclos_sem_peso_dict[ciclo_selecionado]

            print(f"\nAtribuição de peso ao {ciclo_selecionado} realizada com sucesso.")
            break
        else:
            print("\nNúmero de ciclo selecionado inválido. Tente novamente.")

    book.save(caminho_arquivo_excel)


def atribuir_notas(book, turma):
    turma_sheet = book[turma]
    alunos = listar_alunos_em_turma(turma_sheet)
    notas_sheet = book['Notas']

    while True:
        try:

            print("\n\nAlunos disponíveis:")
            for i, aluno in enumerate(alunos, 1):
                print(f"{i}. {aluno}")

            aluno_idx = int(input("\nDigite o número do aluno para atribuir nota (ou 0 para sair): "))
            if aluno_idx == 0:
                break
            elif 1 <= aluno_idx <= len(alunos):
                aluno = alunos[aluno_idx - 1]
                ciclo_atual_idx = int(input(f"Digite o número do ciclo atual: "))
                ciclo_atual_str = f'Ciclo {ciclo_atual_idx}'

            # Verifica se o ciclo já foi atribuído para o aluno específico
                if ciclo_atual_str not in [cell.value for cell in notas_sheet[1]]:
                    notas_sheet.cell(row=1, column=notas_sheet.max_column + 1, value=ciclo_atual_str)
                    ciclo_col = notas_sheet.max_column
                else:
                    ciclo_col = [cell.value for cell in notas_sheet[1]].index(ciclo_atual_str) + 1

                aluno_row = None
                for row in notas_sheet.iter_rows(min_row=2, max_row=notas_sheet.max_row, min_col=1, max_col=1):
                    if row[0].value == aluno:
                        aluno_row = row[0].row
                        break

                if aluno_row is not None and notas_sheet.cell(row=aluno_row, column=ciclo_col).value is not None:
                    print(f"\nO aluno {aluno} já possui uma nota atribuída no {ciclo_atual_str}. Tente novamente.")
                    continue

                nota = float(input(f"Digite a nota para {aluno} no ciclo {ciclo_atual_idx}: "))
                transferir_aluno_para_notas(book, turma, aluno, ciclo_atual_idx, nota)
        except ValueError:
            print("\nErro: tente novamente apenas com números.")
        book.save(caminho_arquivo_excel)


def calcular_medias(book):
    notas_sheet = book['Notas']
    pesos_sheet = book['Pesos']

    # Verificar se existem ciclos e pesos definidos
    if notas_sheet.max_column < 2 or pesos_sheet.max_row < 2:
        print("\nNão há notas ou pesos definidos. Por favor, atribua notas e pesos primeiro.")
        return

    ciclos = [cell.value for cell in notas_sheet[1] if cell.value.startswith('Ciclo ')]
    medias = []

    for aluno_row in notas_sheet.iter_rows(min_row=2, max_row=notas_sheet.max_row, min_col=1, max_col=notas_sheet.max_column):
        aluno = aluno_row[0].value
        notas = aluno_row[1:]
        soma_ponderada = 0.0
        peso_total = 0.0

        for ciclo, nota in zip(ciclos, notas):
            peso = None
            # Encontrar o peso correspondente ao ciclo
            for peso_row in pesos_sheet.iter_rows(min_row=2, max_row=pesos_sheet.max_row, min_col=1, max_col=2):
                if peso_row[0].value == ciclo:
                    peso = peso_row[1].value
                    break

            if peso is not None and nota.value is not None:  # Acesse o valor da célula 'nota'
                soma_ponderada += peso * nota.value  # Acesse o valor da célula 'peso'
                peso_total += peso

        if peso_total > 0:
            media_ponderada = soma_ponderada / peso_total
            medias.append((aluno, media_ponderada))

    # Imprimir as médias ponderadas dos alunos
    print("\nMédias ponderadas dos Alunos:")
    for aluno, media in medias:
        print(f"{aluno}: {media:.2f}")


if __name__ == "__main__":
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

    if os.path.exists(caminho_arquivo_excel):
        book = openpyxl.load_workbook(caminho_arquivo_excel)
    else:
        book = openpyxl.Workbook()

    criar_aba_notas(book)

    print("\nOpções:")
    print("\n1. Atribuir notas")
    print("2. Atribuir pesos")
    print("3. Calcular médias")
    print("4. Sair do programa", "\n")

    opcao = int(input("Escolha uma das opções: "))
    while opcao != 4:
        if opcao == 1:
            print("\n\n\n\n\n\n\n\n")
            turmas = listar_turmas(book)
            turma_escolhida_idx = int(input("\nDigite o número da turma: "))
            turma_escolhida = turmas[turma_escolhida_idx - 1]
            atribuir_notas(book, turma_escolhida)
        elif opcao == 3:
            print("\n\n\n\n\n\n\n\n")
            calcular_medias(book)
        elif opcao == 2:
            atribuir_pesos(book)
        print("\nOpções:")
        print("\n1. Atribuir notas")
        print("2. Atribuir pesos")
        print("3. Calcular médias")
        print("4. Sair do programa", "\n")

        opcao = int(input("Escolha uma das opções: "))
    print("\nEncerrando programa...")
    book.save(caminho_arquivo_excel)