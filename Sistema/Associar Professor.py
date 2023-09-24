import openpyxl
from openpyxl.styles import Alignment, Font

# Função para verificar se um professor já está alocado em alguma turma
def professor_em_turma(planilha, professor_nome):
    abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

    for turma in abas_turmas:
        aba_turma = planilha[turma]
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
            nome = row[0].value
            cpf = row[1].value
            if nome == professor_nome:
                return True

    return False

# Função para verificar se um professor já está em alguma turma
def professor_em_alguma_turma(planilha, professor_nome):
    abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

    for turma in abas_turmas:
        if professor_em_turma(planilha, professor_nome):
            return True

    return False

# Função para adicionar um professor a uma turma específica
def adicionar_professor_a_turma(planilha, turma_destino, professor_nome, professor_cpf):
    aba_turma = planilha[turma_destino]

    # Encontrar a primeira linha vazia após os cabeçalhos "Professores" e "CPF - Prof"
    primeira_linha_vazia = 2  # Começando na linha 2 para evitar os cabeçalhos

    for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
        if all(cell.value is None for cell in row):
            break
        primeira_linha_vazia += 1

    # Adicionar o professor à turma com nome e CPF
    aba_turma.cell(row=primeira_linha_vazia, column=4).value = professor_nome
    aba_turma.cell(row=primeira_linha_vazia, column=5).value = professor_cpf
    print(f"Professor {professor_nome} adicionado à {turma_destino} com sucesso.")
    planilha.save('Dados Cadastrais.xlsx')

# Abrir a planilha
planilha = openpyxl.load_workbook('Dados Cadastrais.xlsx')

while True:
    print("\nOpções:")
    print("\n1. Adicionar professor a uma turma")
    print("2. Mostrar professores disponíveis e não alocados")
    print("3. Sair do programa", "\n")

    escolha = input("Escolha uma das opções: ")

    if escolha == '1':
        # Listar as abas de turma disponíveis
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        if not abas_turmas:
            print("Não foram encontradas abas de turma na planilha.")
        else:
            print("\nTurmas existentes:", "\n")
            for turma in abas_turmas:
                print(turma, "\n")

            # Solicitar ao usuário em qual turma deseja adicionar o professor
            turma_desejada = input("Em qual turma deseja adicionar o professor: ")
            if turma_desejada in abas_turmas:
                aba_turma = planilha[turma_desejada]
                professor_nome = input("Digite o nome do professor: ")
                professor_cpf = input("Digite o CPF do professor: ")

                adicionar_professor_a_turma(planilha, turma_desejada, professor_nome, professor_cpf)
            else:
                print("\nTurma não encontrada.")
    elif escolha == '2':
        print("\nProfessores disponíveis e não alocados:")
        aba_professores = planilha['Professores']

        for row in aba_professores.iter_rows(min_row=2, max_row=aba_professores.max_row, min_col=1, max_col=2):
            nome = row[0].value
            cpf = row[1].value

            if not professor_em_alguma_turma(planilha, nome):
                print(f"Nome: {nome}, CPF: {cpf}")
    elif escolha == '3':
        break
    else:
        print("\nOpção inválida. Escolha 1 para adicionar um professor a uma turma, 2 para mostrar professores disponíveis ou 3 para sair.")

print("\nPrograma encerrado.", "\n")
