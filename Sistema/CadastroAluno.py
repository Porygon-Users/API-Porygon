import openpyxl
import random
import string

# Função para gerar RM e senha aleatórios
def gerar_rm():
    return ''.join(random.choice(string.ascii_uppercase) for _ in range(6))

def gerar_senha():
    return ''.join(random.choice(string.digits) for _ in range(5))

# Função para criar a planilha "dados" se não existir
def criar_planilha_se_necessario():
    try:
        wb = openpyxl.load_workbook('dados.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Alunos"
        
        sheet['A1'] = "RM"
        sheet['B1'] = "Nome Completo"
        sheet['C1'] = "Email"
        sheet['D1'] = "Curso"
        sheet['E1'] = "Senha"
        
        wb.save('dados.xlsx')

# Função para cadastrar um novo aluno
def cadastrar_aluno():
    criar_planilha_se_necessario()
    
    wb = openpyxl.load_workbook('dados.xlsx')
    aluno_sheet = wb['Alunos']
    
    nome = input("\nDigite o seu nome completo: ")
    email = input("Digite o seu email: ")
    curso = input("Digite o seu curso: ")
    senha = gerar_senha()
    
    rm = gerar_rm()
    
    aluno_sheet.append([rm, nome, email, curso, senha])
    
    wb.save('dados.xlsx')
    print(f"Seu RM é: {rm}")
    print(f"Sua senha é: {senha}")
    print("Você foi cadastrado com sucesso!")

# Chamando a função para cadastrar um novo aluno
cadastrar_aluno()
