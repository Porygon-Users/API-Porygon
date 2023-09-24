import openpyxl

# Função para criar grupos em uma turma
def criar_grupos(planilha, turma_nome, num_alunos_por_grupo):
    # Carregar o arquivo da planilha
    wb = openpyxl.load_workbook(planilha)
    
    # Selecionar a aba da turma
    try:
        sheet = wb[turma_nome]
    except KeyError:
        print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
        return
    
    # Obter o índice da coluna que contém o cabeçalho "Grupos"
    coluna_grupos = None
    for cell in sheet[1]:  # Percorre as células da primeira linha
        if cell.value == "Grupos":
            coluna_grupos = cell.column_letter
            break
    
    if coluna_grupos is None:
        print("Não foi encontrada uma coluna com o cabeçalho 'Grupos'.")
        return
    
    # Obter a lista de alunos na turma
    alunos = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row if cell.value]
    
    # Exibir a quantidade de alunos na turma
    num_alunos = len(alunos)
    print(f"\nA turma '{turma_nome}' possui {num_alunos} alunos.")
    
    # Verificar se é possível criar grupos com o número especificado de alunos
    if num_alunos % num_alunos_por_grupo != 0:
        print(f"\nO número de alunos ({num_alunos}) não é divisível pelo número de alunos por grupo ({num_alunos_por_grupo}).")
        return
    
    # Calcular o número total de grupos
    num_grupos = num_alunos // num_alunos_por_grupo
    
    # Criar grupos
    grupos = []
    for i in range(num_grupos):
        grupo_alunos = alunos[i * num_alunos_por_grupo: (i + 1) * num_alunos_por_grupo]
        grupos.append(f'Grupo {i + 1} - {", ".join(grupo_alunos)}')
    
    # Adicionar os grupos à coluna "Grupos" na planilha
    for i, grupo in enumerate(grupos):
        sheet[f"{coluna_grupos}{i + 2}"] = grupo
    
    # Salvar a planilha
    wb.save(planilha)
    print(f"\nForam criados {num_grupos} grupos com {num_alunos_por_grupo} alunos cada na '{turma_nome}'.")

# Função para contar grupos em uma turma
def contar_grupos(planilha, turma_nome):
    # Carregar o arquivo da planilha
    wb = openpyxl.load_workbook(planilha)
    
    # Verificar se a turma existe
    if turma_nome not in wb.sheetnames:
        print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
        return

    # Selecionar a aba da turma
    sheet = wb[turma_nome]

    # Obter o índice da coluna que contém o cabeçalho "Grupos"
    coluna_grupos = None
    for cell in sheet[1]:  # Percorre as células da primeira linha
        if cell.value == "Grupos":
            coluna_grupos = cell.column_letter
            break
    
    if coluna_grupos is None:
        print("Não foi encontrada uma coluna com o cabeçalho 'Grupos'.")
        return

    # Contar grupos na coluna "Grupos" da turma
    num_grupos = sum(1 for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=coluna_grupos, max_col=coluna_grupos) if row[0].value)
    
    if num_grupos > 0:
        print(f"\nA turma '{turma_nome}' possui {num_grupos} grupos.")
        alunos_por_grupo = len(sheet.cell(row=2, column=coluna_grupos).value.split(','))  # Assumindo que todos os grupos têm o mesmo número de alunos
        print(f"\nCada grupo possui {alunos_por_grupo} alunos.")
    else:
        print(f"\nA turma '{turma_nome}' não possui grupos.")

# Função para contar alunos em uma turma
def contar_alunos(planilha, turma_nome):
    # Carregar o arquivo da planilha
    wb = openpyxl.load_workbook(planilha)
    
    # Selecionar a aba da turma
    try:
        sheet = wb[turma_nome]
    except KeyError:
        print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
        return
    
    # Contar alunos na turma
    num_alunos = sum(1 for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) if row[0].value)
    print(f"\nA '{turma_nome}' possui {num_alunos} alunos.")
    return num_alunos  # Retornar a quantidade de alunos na turma

# Função para listar as turmas existentes
def listar_turmas(planilha):
    wb = openpyxl.load_workbook(planilha)
    turmas = [sheet for sheet in wb.sheetnames if sheet.startswith('Turma ')]
    if turmas:
        print("\nTurmas existentes:")
        for turma in turmas:
            print(turma)
    else:
        print("\nNão foram encontradas abas de turma na planilha.")

# Função principal do programa
def main():
    planilha = 'Dados Cadastrais.xlsx'
    
    while True:
        print("\nOpções:")
        print("\n1. Criar grupos")
        print("2. Ver quantos grupos têm em uma turma")
        print("3. Listar turmas existentes")
        print("4. Sair", "\n")
        
        escolha = input("Escolha uma opção: ")
        
        if escolha == '1':
            turma = input('\nDigite o nome da turma: ')
            alunos_na_turma = contar_alunos(planilha, turma)  # Mostrar a quantidade de alunos na turma
            if alunos_na_turma is not None:
                if alunos_na_turma > 0:
                    alunos_por_grupo = int(input('Digite o número de alunos por grupo: '))
                    criar_grupos(planilha, turma, alunos_por_grupo)
        elif escolha == '2':
            turma = input('Digite o nome da turma que deseja ver: ')
            contar_grupos(planilha, turma)
        elif escolha == '3':
            listar_turmas(planilha)
        elif escolha == '4':
            print("\nSaindo do programa.", "\n")
            break
        else:
            print("\nOpção inválida. Tente novamente.", "\n")

if __name__ == "__main__":
    main()
