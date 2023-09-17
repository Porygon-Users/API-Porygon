import openpyxl

# Função para consultar notas e frequências do aluno
def consultar_notas_frequencia_aluno():
    wb = openpyxl.load_workbook('dados.xlsx')
    
    notas_sheet_name = "Notas"
    
    if notas_sheet_name not in wb.sheetnames:
        print(f"A aba '{notas_sheet_name}' não foi encontrada.")
        return
    
    notas_sheet = wb[notas_sheet_name]
    
    rm_aluno = input("Digite o RM do aluno: ")
    senha_aluno = input("Digite a senha do aluno: ")
    
    aluno_encontrado = False
    
    # Verifica se o RM e a senha do aluno estão na aba "Alunos"
    alunos_sheet = wb["Alunos"]
    
    for row in alunos_sheet.iter_rows(values_only=True):
        if row[0] == rm_aluno and row[4] == senha_aluno:
            aluno_encontrado = True
            break
    
    if aluno_encontrado:
        rm_na_notas = False
        # Verifica se o RM do aluno está na aba "Notas"
        for row in notas_sheet.iter_rows(values_only=True):
            if row[0] == rm_aluno:
                rm_na_notas = True
                print("\nNotas e Frequência do Aluno:")
                print(f"RM: {row[0]}")
                print(f"Notas: {row[1]}")
                print(f"Faltas: {row[2]}")
                print(f"Presenças: {row[3]}")
                print(f"Prazos de Entrega: {row[4]}")
                break
        if not rm_na_notas:
            print(f"RM {rm_aluno} não encontrado na aba 'Notas'.")
    else:
        print("RM não encontrado na aba 'Alunos' ou senha incorreta.")

# Chamando a função para consultar notas e frequência do aluno
consultar_notas_frequencia_aluno()
