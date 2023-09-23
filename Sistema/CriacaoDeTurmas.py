import openpyxl
import os

# Função para adicionar uma turma à planilha
def criar_ou_carregar_planilha(nome_arquivo):
    if os.path.exists(nome_arquivo):
        print(f"Carregando a planilha '{nome_arquivo}'...")
        return openpyxl.load_workbook(nome_arquivo)
    else:
        print(f"Criando uma nova planilha chamada '{nome_arquivo}'...")
        return openpyxl.Workbook()

# Função para obter o nome do professor
def obter_nome_professor():
    nome_professor = input("Digite o nome do professor: ")
    return nome_professor

# Função para criar uma nova turma
def criar_turma(planilha, turma_nome, nome_professor):
    if turma_nome in planilha.sheetnames:
        print(f"A turma '{turma_nome}' já existe.")
    else:
        planilha.create_sheet(title=turma_nome)
        print(f"Turma '{turma_nome}' criada com sucesso.")
        sheet = planilha[turma_nome]
        sheet['A1'] = "Número"
        sheet['B1'] = "Aluno"
        sheet['C1'] = "Professor"
        sheet['D1'] = nome_professor

# Função para adicionar aluno a uma turma
def adicionar_aluno(planilha, turma_nome, aluno_nome):
    if turma_nome not in planilha.sheetnames:
        print(f"A turma '{turma_nome}' não existe.")
    else:
        sheet = planilha[turma_nome]
        # Obtemos a última linha da coluna A
        linha_aluno = sheet.max_row + 1 if sheet.max_row > 1 else 2
        # Adiciona o número do aluno na coluna A (alunos)
        numero_aluno = linha_aluno - 1  # Número do aluno (começando de 1)
        sheet[f'A{linha_aluno}'] = numero_aluno
        # Adiciona o aluno na coluna B (alunos)
        sheet[f'B{linha_aluno}'] = aluno_nome

def criar_grupo(planilha, turma_nome, grupo_nome):
    if turma_nome not in planilha.sheetnames:
        print(f"A turma '{turma_nome}' não existe.")
    else:
        turma_sheet = planilha[turma_nome]
        grupo_sheet_name = f'Grupos_{turma_nome}'

        # Verificar se a página de grupos já existe
        if grupo_sheet_name in planilha.sheetnames:
            grupo_sheet = planilha[grupo_sheet_name]
        else:
            grupo_sheet = planilha.create_sheet(title=grupo_sheet_name)
            print(f"Página de grupos criada para a turma '{turma_nome}'.")

        # Verificar se o grupo já existe na página de grupos
        grupos_existentes = grupo_sheet.cell(row=1, column=1).value
        if grupos_existentes:
            grupos = grupos_existentes.split(', ')
            if grupo_nome in grupos:
                print(f"O grupo '{grupo_nome}' já existe na turma '{turma_nome}'.")
                return
            else:
                grupos.append(grupo_nome)
                novo_grupos = ', '.join(grupos)
                grupo_sheet.cell(row=1, column=1, value=novo_grupos)
        else:
            grupo_sheet.cell(row=1, column=1, value=grupo_nome)

        # Calcular a posição para a nova coluna
        nova_coluna = len(grupos_existentes.split(', ')) if grupos_existentes else 1
        coluna_letra = chr(65 + (nova_coluna * 2))  # Incrementa o deslocamento a cada novo grupo

        # Adicionar o nome do grupo na célula correspondente
        grupo_sheet[f'{coluna_letra}1'] = grupo_nome
        print(f"Grupo '{grupo_nome}' criado com sucesso na turma '{turma_nome}'.")

        if 'Sheet' in planilha.sheetnames:
            default_sheet = planilha['Sheet']
            planilha.remove(default_sheet)

# Dicionário para rastrear a última linha de cada grupo
ultima_linha_por_grupo = {}

# Função para verificar se o aluno está na turma
def verificar_aluno_na_turma(planilha, turma_nome, aluno_nome):
    if turma_nome not in planilha.sheetnames:
        print(f"A turma '{turma_nome}' não existe.")
        return False
    else:
        turma_sheet = planilha[turma_nome]
        for row in turma_sheet.iter_rows(min_row=2, max_col=2, max_row=turma_sheet.max_row):
            for cell in row:
                if cell.value == aluno_nome:
                    return True
        print(f"O aluno '{aluno_nome}' não existe na turma '{turma_nome}'.")
        return False

# Função para adicionar aluno a um grupo de uma turma
def adicionar_aluno_grupo(planilha, turma_nome, grupo_nome):
    global ultima_linha_por_grupo  # Acessando a variável global
    grupo_sheet_name = f'Grupos_{turma_nome}'

    if grupo_sheet_name not in planilha.sheetnames:
        print(f"O grupo '{grupo_nome}' não existe na turma '{turma_nome}'.")
    else:
        grupo_sheet = planilha[grupo_sheet_name]

        # Encontre a coluna correspondente ao grupo
        coluna_grupo = None
        for col_idx in range(1, grupo_sheet.max_column + 1, 2):  # Pula 1 coluna para cada grupo criado
            if grupo_sheet.cell(row=1, column=col_idx).value == grupo_nome:
                coluna_grupo = col_idx
                break

        if coluna_grupo is not None:
            while True:
                aluno_nome = input("Digite o nome do aluno (ou 's' para sair): ")

                if aluno_nome.lower() == 's':
                    break

                # Verifique se o aluno está na turma
                if verificar_aluno_na_turma(planilha, turma_nome, aluno_nome):
                    # Obter a última linha do grupo
                    ultima_linha = ultima_linha_por_grupo.get(grupo_nome, 1)

                    # Encontrar a próxima linha vazia abaixo do grupo
                    proxima_linha = ultima_linha + 1
                    while grupo_sheet.cell(row=proxima_linha, column=coluna_grupo).value:
                        proxima_linha += 1

                    # Adicione o aluno abaixo do nome do grupo
                    grupo_sheet.cell(row=proxima_linha, column=coluna_grupo, value=aluno_nome)
                    ultima_linha_por_grupo[grupo_nome] = proxima_linha  # Atualize a última linha
                    print(f"Aluno '{aluno_nome}' adicionado ao grupo '{grupo_nome}' na turma '{turma_nome}'.")
        else:
            print(f"Grupo '{grupo_nome}' não encontrado na turma '{turma_nome}'.")

# Nome do arquivo para a planilha
nome_arquivo = "cadastro_turmas.xlsx"

# Criar ou carregar a planilha
planilha = criar_ou_carregar_planilha(nome_arquivo)

while True:
    print("\n\nEscolha uma opção:")
    print("1. Criar turma")
    print("2. Visualizar turmas")
    print("3. Adicionar aluno a turma")
    print("4. Criar grupo em uma turma")
    print("5. Adicionar aluno a um grupo de uma turma")
    print("6. Adicionar grupo a uma turma existente")
    print("7. Sair")
    escolha = input("Opção: ")
    print("\n\n")

    if escolha == '1':
        while True:
            nome_turma = input("Digite o nome da turma (ou 's' para sair) : ").lower()
        
            if nome_turma == "s":
                break
        
            nome_professor = obter_nome_professor()  # Obter o nome do professor
            criar_turma(planilha, nome_turma, nome_professor)

    elif escolha == '2':
        print("Turmas disponíveis:")
        for turma in planilha.sheetnames:
            print(turma)
    
    elif escolha == '3':
        while True:
            nome_turma = input("Digite o nome da turma em que deseja adicionar os alunos (ou 's' para sair): ").lower()

            if nome_turma.lower() == 's':
                break

            if nome_turma not in planilha.sheetnames:
                print(f"A turma '{nome_turma}' não existe. Tente novamente.")
            else:
                while True:
                    aluno_nome = input("Digite o nome do aluno: ")
                    adicionar_aluno(planilha, nome_turma, aluno_nome)

                    while True:
                        continuar = input("Deseja adicionar outro aluno? (s/n): ").lower()
                        if continuar in ('s', 'n'):
                            break
                        else:
                            print("Opção inválida. Por favor, digite 's' para sim ou 'n' para não.")

                    if continuar == 'n':
                        break
    
    elif escolha == '4':
        while True:
            nome_turma = input("Digite o nome da turma (ou 's' para sair): ")
            
            if nome_turma == 's':
                break
            
            nome_grupos = input("Digite os nomes dos grupos separados por vírgula: ")
            criar_grupo(planilha, nome_turma, nome_grupos)

    elif escolha == '5':
        while True:
            nome_turma = input("Digite o nome da turma (ou 's' para sair): ")
            
            if nome_turma == 's':
                break
            
            nome_grupo = input("Digite o nome do grupo: ")
            adicionar_aluno_grupo(planilha, nome_turma, nome_grupo)

    elif escolha == '6':
        while True:
            nome_turma = input("Digite o nome da turma em que deseja adicionar o grupo (ou 's' para sair): ").lower()
            if nome_turma.lower() == 's':
                break
            if nome_turma not in planilha.sheetnames:
                print(f"A turma '{nome_turma}' não existe. Tente novamente.")
            else:
                nome_grupo = input("Digite o nome do grupo: ")
                criar_grupo(planilha, nome_turma, nome_grupo)

    elif escolha == '7':
        planilha.save(nome_arquivo)
        print(f"Planilha salva como '{nome_arquivo}'.")
        break
    else:
        print("Opção inválida. Tente novamente.")
