import openpyxl
from datetime import datetime, timedelta

# Função para adicionar a data de início do semestre e calcular 180 dias após
def adicionar_data_inicio_e_calcula_fim(planilha, turma_destino, data_inicio_semestre):
    aba_turma = planilha[turma_destino]

    # Converter a data de início para um objeto datetime
    data_inicio = datetime.strptime(data_inicio_semestre, "%d/%m/%Y")

    # Calcular a data 180 dias após a data de início
    data_fim = data_inicio + timedelta(days=180)

    # Adicionar a data de início na coluna F
    for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=6, max_col=6):
        row[0].value = data_inicio.strftime("%d/%m/%Y")

    # Adicionar a data de fim na coluna G
    for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=7, max_col=7):
        row[0].value = data_fim.strftime("%d/%m/%Y")

    planilha.save('Dados Cadastrais.xlsx')

# Abrir a planilha
planilha = openpyxl.load_workbook('Dados Cadastrais.xlsx')

while True:
    print("\nOpções:")
    print("1. Adicionar data de início do semestre")
    print("2. Sair do programa")

    escolha = input("Escolha a opção (1 ou 2): ")

    if escolha == '1':
        # Listar as abas de turma disponíveis
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        if not abas_turmas:
            print("Não foram encontradas abas de turma na planilha.")
        else:
            print("Turmas existentes:")
            for i, turma in enumerate(abas_turmas, start=1):
                print(f"{i}. {turma}")

            # Solicitar ao usuário qual turma deseja adicionar a data
            num_turma = int(input("Escolha o número da turma: "))
            if 1 <= num_turma <= len(abas_turmas):
                turma_desejada = abas_turmas[num_turma - 1]
                data_inicio_semestre = input("Digite a data de início do semestre (DD/MM/AAAA): ")
                adicionar_data_inicio_e_calcula_fim(planilha, turma_desejada, data_inicio_semestre)
                print("Data de início e fim do semestre adicionadas com sucesso.")
            else:
                print("Número de turma inválido.")
    elif escolha == '2':
        break
    else:
        print("Opção inválida. Escolha 1 para adicionar a data de início do semestre e calcular a data de fim ou 2 para sair.")

print("Programa encerrado.")
