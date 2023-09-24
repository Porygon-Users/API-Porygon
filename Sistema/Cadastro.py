import os
import random
import openpyxl
from openpyxl import Workbook

# Gera uma senha aleatória com letras maiúsculas, minúsculas e números, contendo no máximo 7 caracteres
def gerar_senha():
    caracteres = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    senha = ''.join(random.choice(caracteres) for i in range(7))
    return senha

# Validação de CPF como números com 11 caracteres
def validar_cpf(cpf):
    if cpf.isdigit() and len(cpf) == 11:
        return True
    else:
        return False

# Formata o CPF com os "-" nos dois últimos dígitos
def formatar_cpf(cpf):
    return f"{cpf[:9]}-{cpf[9:]}"

# Verifica se já existe um banco de dados, caso não, ele cria
arquivo_excel = "Dados Cadastrais.xlsx"
if not os.path.exists(arquivo_excel):
    workbook = Workbook()
    workbook.remove(workbook["Sheet"])  # Remove a aba "Sheet" padrão
    workbook.create_sheet("Alunos")  # Cria a aba para alunos
    workbook.create_sheet("Professores")  # Cria a aba para professores
    sheet_alunos = workbook["Alunos"]
    sheet_alunos.append(["Nome", "CPF", "Email", "Login", "Senha"])
    sheet_professores = workbook["Professores"]
    sheet_professores.append(["Nome", "CPF", "Email", "Login", "Senha"])
    workbook.save(arquivo_excel)
else:
    workbook = openpyxl.load_workbook(arquivo_excel)
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])  # Remove a aba "Sheet" padrão
    sheet_alunos = workbook["Alunos"]
    sheet_professores = workbook["Professores"]

while True:
    tipo_cadastro = input("\nDigite A para aluno ou P para professor: ").strip().upper()

    if tipo_cadastro not in ('A', 'P'):
        print("Por favor, tente novamente e digite 'A' para aluno ou 'P' para professor.")
        continue
    
    nome = input("\nDigite o nome: ")
    
    # Validação de CPF
    while True:
        cpf = input("Digite o CPF: ")
        if validar_cpf(cpf):
            cpf = formatar_cpf(cpf)  # Formata o CPF com os "-"
            break
        else:
            print("\nPor favor, tente novamente.")
    
    email = input("Digite o email: ")

    login = nome.lower() + str(random.randint(10, 99))
    
    if tipo_cadastro == 'A':
        senha = gerar_senha()
    else:
        senha = cpf  # Senha para professores é o próprio CPF
    
    if tipo_cadastro == 'A':
        sheet_alunos.append([nome, cpf, email, login, senha])
    else:
        sheet_professores.append([nome, cpf, email, login, senha])
    
    workbook.save(arquivo_excel)

    print("\n\nCadastro concluído.")
    print(f"Login: {login}")
    print(f"Senha: {senha}")

    while True:
        resposta = input("\nDeseja cadastrar outro aluno/professor? (S para continuar ou N para sair): ").strip().upper()
        if resposta == 'S':
            break
        elif resposta == 'N':
            break
        else:
            print("Por favor, tente novamente e responda com 'S' para continuar ou 'N' sair.")
    
    if resposta == 'N':
        break

print("\n")
