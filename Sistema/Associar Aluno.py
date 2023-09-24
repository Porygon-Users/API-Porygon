import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# Função para verificar se um aluno já está alocado em alguma turma
def aluno_em_turma(planilha, aluno_chave):
    abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]
    
    for turma in abas_turmas:
        aba_turma = planilha[turma]
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
            nome = row[0].value
            cpf = row[1].value
            aluno_turma_chave = (nome, cpf)
            
            if aluno_turma_chave == aluno_chave:
                return True
    
    return False

# Função para adicionar alunos à turma
def adicionar_alunos_a_turma(planilha, turma_destino, alunos_adicionados, quantidade):
    aba_turma = planilha[turma_destino]

    # Encontrar a primeira linha vazia após os cabeçalhos "Alunos", "CPF" e "Email"
    primeira_linha_vazia = 2  # Começando na linha 2 para evitar os cabeçalhos
    
    for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
        if all(cell.value is None for cell in row):
            break
        primeira_linha_vazia += 1
    
    if quantidade <= 0:
        print("Quantidade inválida. A quantidade deve ser maior que zero.")
        return
    
    for _ in range(quantidade):
        nome = f"Aluno {primeira_linha_vazia}"
        cpf = f"CPF {primeira_linha_vazia}"
        email = f"Email{primeira_linha_vazia}@exemplo.com"
        aluno_chave = (nome, cpf)

        # Verificar se o aluno já está alocado em alguma turma
        if aluno_em_turma(planilha, aluno_chave):
            print(f"O aluno {nome} com CPF {cpf} já está alocado em outra turma.")
            continue

        # Verificar se o aluno já foi adicionado a esta turma
        aluno_completo = (nome, cpf, email)
        if aluno_completo not in alunos_adicionados:
            # Adicionar o aluno à turma com nome, CPF e email
            nova_linha = [nome, cpf, email]
            aba_turma.append(nova_linha)
            alunos_adicionados.add(aluno_completo)
            print(f"Aluno {nome} adicionado à {turma_destino} com sucesso.")
            primeira_linha_vazia += 1

    planilha.save('Dados Cadastrais.xlsx')

# Abrir a planilha
planilha = openpyxl.load_workbook('Dados Cadastrais.xlsx')

# Conjunto para manter o controle dos alunos já adicionados
alunos_adicionados = set()

while True:
    print("\nOpções:")
    print("\n1. Adicionar alunos às turmas")
    print("2. Ver alunos disponíveis e não alocados")
    print("3. Sair do programa", "\n")
    
    escolha = input("Escolha uma das opções: ")
    
    if escolha == '1':
        # Listar as abas de turma disponíveis
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]
        
        if not abas_turmas:
            print("\nNão foram encontradas abas de turma na planilha.", "\n")
        else:
            print("\nTurmas existentes:", "\n")
            for turma in abas_turmas:
                print(turma, "\n")
            
            # Solicitar ao usuário em qual turma deseja adicionar os alunos
            turma_desejada = input("Em qual turma deseja adicionar os alunos: ")
            if turma_desejada in abas_turmas:
                quantidade_alunos = int(input("Quantos alunos deseja adicionar: "))
                adicionar_alunos_a_turma(planilha, turma_desejada, alunos_adicionados, quantidade_alunos)
            else:
                print("\nTurma não encontrada.")
    elif escolha == '2':
        alunos_nao_alocados = 0
        
        # Verificar alunos na aba "Alunos" que não foram alocados a nenhuma turma
        aba_alunos = planilha['Alunos']
        for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=3):
            nome = row[0].value
            cpf = row[1].value
            email = row[2].value
            aluno_chave = (nome, cpf)
            
            if not aluno_em_turma(planilha, aluno_chave):
                alunos_nao_alocados += 1
                
        print(f"\nAlunos disponíveis para alocação: {alunos_nao_alocados}")
    elif escolha == '3':
        break
    else:
        print("\nOpção inválida. Escolha 1 para adicionar alunos, 2 para verificar alunos disponíveis ou 3 para sair.")

print("\nPrograma encerrado.", "\n")
