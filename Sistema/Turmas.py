import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# Nome do arquivo Excel
arquivo_excel = "Dados Cadastrais.xlsx"

# Carregar o arquivo Excel (se existir) ou criar um novo
try:
    book = openpyxl.load_workbook(arquivo_excel)
except FileNotFoundError:
    book = openpyxl.Workbook()
    book.save(arquivo_excel)

# Função para criar uma nova aba de turma
def criar_nova_turma(book, numero_turma):
    nova_aba = book.create_sheet(f"Turma {numero_turma}")
    cabecalhos = ["Alunos", "CPF", "Email", "Professores", "CPF - Prof", "Grupos", "Inicio de Semestre", "Fim de Semestre"]
    nova_aba.append(cabecalhos)

    # Ajustar a largura da coluna e centralizar os cabeçalhos
    for col_idx, header in enumerate(cabecalhos, 1):
        coluna_letra = get_column_letter(col_idx)
        cell = nova_aba[f"{coluna_letra}1"]
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        cell.value = header
        nova_aba.column_dimensions[coluna_letra].width = 35

    book.save(arquivo_excel)

# Função para exibir o número de turmas
def mostrar_numero_de_turmas(book):
    turmas_existentes = [sheet for sheet in book.sheetnames if sheet.startswith("Turma ")]
    if not turmas_existentes:
        print("\nNenhuma turma encontrada.")
    else:
        print("\nTurmas existentes:")
        for turma in turmas_existentes:
            print("\n", turma)
        print(f"\nTotal de turmas: {len(turmas_existentes)}")

# Função para excluir uma ou mais turmas
def excluir_turmas(book, turmas_a_excluir):
    for turma_nome in turmas_a_excluir:
        if turma_nome in book.sheetnames:
            book.remove(book[turma_nome])
            print(f"\nTurma {turma_nome} excluída com sucesso.")
        else:
            print(f"\nA turma {turma_nome} não foi encontrada.")
    book.save(arquivo_excel)

    # Loop principal
while True:
    print("\nEscolha uma opção:")
    print("\n1 - Criar nova turma")
    print("2 - Visualizar turmas")
    print("3 - Excluir turmas")
    print("4 - Sair", "\n")
    
    opcao = input('Digite o número da opção: ')
    
    if opcao == "1":
        quantidade_turmas = int(input("Digite a quantidade de turmas que deseja criar: "))
        for _ in range(quantidade_turmas):
            proxima_turma = sum(1 for sheet in book.sheetnames if sheet.startswith("Turma ")) + 1
            criar_nova_turma(book, proxima_turma)
            print(f"\nTurma {proxima_turma} criada com sucesso.")
    elif opcao == "2":
        mostrar_numero_de_turmas(book)
    elif opcao == "3":
        turmas_a_excluir = input("Digite o nome das turmas que deseja excluir: ")
        turmas_a_excluir = [turma.strip() for turma in turmas_a_excluir.split(",")]
        excluir_turmas(book, turmas_a_excluir)
    elif opcao == "4":
        break
    else:
        print("\nOpção inválida.", "\n")

print("\nEncerrando o programa.", "\n")