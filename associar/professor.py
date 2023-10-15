import openpyxl
import os

# Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__))

# Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

# Abrir o arquivo Excel existente ou criar um novo
if os.path.exists(caminho_arquivo_excel):
    book = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    book = openpyxl.Workbook()

# Função para verificar se um professor já está alocado em uma turma
def professor_em_turma(planilha, professor_chave):
    abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

    for turma in abas_turmas:
        aba_turma = planilha[turma]
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
            nome = row[0].value
            cpf = row[1].value
            professor_turma_chave = (nome, cpf)

            if professor_turma_chave == professor_chave:
                return turma  # Retorna o nome da turma onde o professor está

    return None

# Função para listar os professores em uma turma
def listar_professores_na_turma(planilha, turma_nome):
    if turma_nome in planilha.sheetnames:
        aba_turma = planilha[turma_nome]
        professores_na_turma = []
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
            nome = row[0].value
            cpf = row[1].value
            professores_na_turma.append((nome, cpf))
        return professores_na_turma
    else:
        return None

# Conjunto para manter o controle dos professores já adicionados
professores_adicionados = set()

while True:
    print("\nOpções:")
    print("\n1. Adicionar professores às turmas")
    print("2. Ver professores disponíveis e não alocados")
    print("3. Remover professor de uma turma")
    print("4. Sair do programa", "\n")

    escolha = input("Escolha uma das opções: ")

    if escolha == '1':
        # Listar as abas de turma disponíveis (apenas as abertas)
        abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ')]

        if not abas_turmas:
            print("\nNão foram encontradas abas de turma na planilha.", "\n")
        else:
            print("\nTurmas abertas:")
            for i, turma in enumerate(abas_turmas, start=1):
                if "(fechada)" not in turma:
                    print(f"{turma.replace('(fechada)', '')}")

            # Pergunta ao usuário em qual turma deseja adicionar os professores
            try:
                indice_turma = int(input("Digite o número da turma que deseja adicionar os professores: ")) - 1

                if 0 <= indice_turma < len(abas_turmas) and "(fechada)" not in abas_turmas[indice_turma]:
                    turma_desejada = abas_turmas[indice_turma]
                    quantidade_professores = int(input("Quantos professores deseja adicionar: "))

                    # Verificar professores na aba "Cadastro" que não foram colocados em nenhuma turma
                    aba_professores = book['Cadastro']
                    professores_disponiveis = []

                    for row in aba_professores.iter_rows(min_row=2, max_row=aba_professores.max_row, min_col=1, max_col=4):
                        nome = row[0].value
                        cpf = row[1].value
                        funcao = row[3].value
                        professor_chave = (nome, cpf)

                        if not professor_em_turma(book, professor_chave) and funcao == "professor":
                            professores_disponiveis.append((nome, cpf))

                    # Selecionar a quantidade desejada de professores disponíveis
                    professores_selecionados = professores_disponiveis[:quantidade_professores]

                    # Função para adicionar um professor a uma turma específica
                    def adicionar_professor_a_turma(planilha, turma_destino, professor_nome, professor_cpf):
                        aba_turma = planilha[turma_destino]

                        # Encontrar a primeira linha vazia após os cabeçalhos "Professores" e "CPF - Prof"
                        primeira_linha_vazia = 2  # Começando na linha 2 para evitar os cabeçalhos

                        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
                            if row[0].value is None or row[1].value is None:
                                break
                            primeira_linha_vazia += 1


                        # Agora, você pode adicionar os dados do professor na primeira linha vazia
                        aba_turma.cell(row=primeira_linha_vazia, column=4).value = professor_nome
                        aba_turma.cell(row=primeira_linha_vazia, column=5).value = professor_cpf

                        # Salvar o livro após adicionar o professor
                        planilha.save(caminho_arquivo_excel)

                    # Adicionar os professores selecionados à turma
                    aba_turma = book[turma_desejada]
                    for professor in professores_selecionados:
                        nome, cpf = professor
                        adicionar_professor_a_turma(book, turma_desejada, nome, cpf)
                        print(f"Professor {nome} adicionado à {turma_desejada} com sucesso.")
                        professores_adicionados.add(professor)

                    else:
                        print("\nTurma não encontrada ou está fechada.")
            except ValueError:
                print("\nDigite um número válido para selecionar a turma.")

    elif escolha == '2':
        professores_nao_alocados = 0

        # Verificar professores na aba "Cadastro" que não foram alocados a nenhuma turma
        aba_professores = book['Cadastro']
        for row in aba_professores.iter_rows(min_row=2, max_row=aba_professores.max_row, min_col=1, max_col=4):
            nome = row[0].value
            cpf = row[1].value
            funcao = row[3].value
            professor_chave = (nome, cpf)

            if not professor_em_turma(book, professor_chave) and funcao == "professor":
                professores_nao_alocados += 1

        print(f"\nProfessores disponíveis para alocação: {professores_nao_alocados}")

    elif escolha == '3':
        # Remover professor de uma turma
        turma_escolhida = input("Digite o nome da turma da qual deseja remover o professor: ")
        professores_na_turma = listar_professores_na_turma(book, turma_escolhida)

        if professores_na_turma:
            print(f"Professores na turma {turma_escolhida}:")
            for i, professor in enumerate(professores_na_turma, start=1):
                nome, cpf = professor
                print(f"{i}. {nome} (CPF: {cpf})")

            escolha_professor = input("Digite o número do professor que deseja remover: ")

            try:
                escolha_professor = int(escolha_professor)
                if 1 <= escolha_professor <= len(professores_na_turma):
                    professor_remover = professores_na_turma[escolha_professor - 1]
                    nome_professor, cpf_professor = professor_remover

                    turma_do_professor = professor_em_turma(book, (nome_professor, cpf_professor))
                    if turma_do_professor:
                        aba_turma = book[turma_do_professor]
                        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
                            nome = row[0].value
                            cpf = row[1].value
                            if (nome, cpf) == (nome_professor, cpf_professor):
                                aba_turma.delete_rows(row[0].row)
                                print(f"Professor {nome_professor} removido da turma {turma_do_professor}.")
                                break
                    else:
                        print("Professor não encontrado em nenhuma turma.")
            except ValueError:
                print("Escolha de professor inválida.")

    elif escolha == '4':
        break
    else:
        print("\nOpção inválida. Escolha 1 para adicionar professores, 2 para verificar professores disponíveis, 3 para remover um professor de uma turma, ou 4 para sair.")

print("\nAlterações realizadas com sucesso, encerrando o programa.", "\n")

# Salve as alterações no arquivo Excel
book.save(caminho_arquivo_excel)