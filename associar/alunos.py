import openpyxl
import os

# Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__ ))

# Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

# Abrir o arquivo Excel existente ou criar um novo
if os.path.exists(caminho_arquivo_excel):
    book = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    book = openpyxl.Workbook()

# Função para verificar se um aluno já está alocado em uma turma
def aluno_em_turma(planilha, aluno_chave):
    abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

    for turma in abas_turmas:
        aba_turma = planilha[turma]
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
            nome = row[0].value
            cpf = row[1].value
            aluno_turma_chave = (nome, cpf)

            if aluno_turma_chave == aluno_chave:
                return turma  # Retorna o nome da turma onde o aluno está

    return None

# Função para listar os alunos em uma turma
def listar_alunos_na_turma(planilha, turma_nome):
    if turma_nome in planilha.sheetnames:
        aba_turma = planilha[turma_nome]
        alunos_na_turma = []
        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
            nome = row[0].value
            cpf = row[1].value
            alunos_na_turma.append((nome, cpf))
        return alunos_na_turma
    else:
        return None

# Conjunto para manter o controle dos alunos já adicionados
alunos_adicionados = set()

while True:
    print("\nOpções:")
    print("\n1. Adicionar alunos às turmas")
    print("2. Ver alunos disponíveis e não alocados")
    print("3. Remover aluno de uma turma")
    print("4. Sair do programa", "\n")

    escolha = input("Escolha uma das opções: ")

    if escolha == '1':
        #Listar as abas de turma disponíveis (apenas as abertas)
        abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ')]

        if not abas_turmas:
            print("\nNão foram encontradas abas de turma na planilha.", "\n")
        else:
            print("\nTurmas abertas:")
            for i, turma in enumerate(abas_turmas, start=1):
                if "(fechada)" not in turma:
                    print(f"{turma.replace('(fechada)', '')}")

            #Pergunta ao usuário em qual turma deseja adicionar os alunos
            try:
                indice_turma = int(input("Digite o número da turma que deseja adicionar os alunos: ")) - 1

                if 0 <= indice_turma < len(abas_turmas) and "(fechada)" not in abas_turmas[indice_turma]:
                    turma_desejada = abas_turmas[indice_turma]
                    quantidade_alunos = int(input("Quantos alunos deseja adicionar: "))

                    #Verificar alunos na aba "Alunos" que não foram colocados em nenhuma turma
                    aba_alunos = book['Cadastro']
                    alunos_disponiveis = []

                    for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=4):
                        nome = row[0].value
                        cpf = row[1].value
                        funcao = row[3].value
                        aluno_chave = (nome, cpf)

                        if not aluno_em_turma(book, aluno_chave) and funcao == "aluno":
                            alunos_disponiveis.append((nome, cpf, row[2].value))

                    #Selecionar a quantidade desejada de alunos disponíveis
                    alunos_selecionados = alunos_disponiveis[:quantidade_alunos]

                    # Adicionar os alunos selecionados à turma
                    aba_turma = book[turma_desejada]
                    for aluno in alunos_selecionados:
                        nome, cpf, email = aluno
                        nova_linha = [nome, cpf, email]
                        linha = [nome, cpf, email]
                        aba_turma.append(linha)
                        print(f"Aluno {nome} adicionado à {turma_desejada} com sucesso.")
                        alunos_adicionados.add(aluno)

                else:
                    print("\nTurma não encontrada ou está fechada.")
            except ValueError:
                print("\nDigite um número válido para selecionar a turma.")

    elif escolha == '2':
        alunos_nao_alocados = 0

        #Verificar alunos na aba "alunos" que não foram alocados a nenhuma turma
        aba_alunos = book['Cadastro']
        for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=4):
            nome = row[0].value
            cpf = row[1].value
            funcao = row[3].value
            aluno_chave = (nome, cpf)

            if not aluno_em_turma(book, aluno_chave) and funcao == "aluno":
                alunos_nao_alocados += 1

        print(f"\nAlunos disponíveis para alocação: {alunos_nao_alocados}")

    elif escolha == '3':
        # Remover aluno de uma turma
        turma_escolhida = input("Digite o nome da turma da qual deseja remover o aluno: ")
        alunos_na_turma = listar_alunos_na_turma(book, turma_escolhida)

        if alunos_na_turma:
            print(f"Alunos na turma {turma_escolhida}:")
            for i, aluno in enumerate(alunos_na_turma, start=1):
                nome, cpf = aluno
                print(f"{i}. {nome} (CPF: {cpf})")

            escolha_aluno = input("Digite o número do aluno que deseja remover: ")

            try:
                escolha_aluno = int(escolha_aluno)
                if 1 <= escolha_aluno <= len(alunos_na_turma):
                    aluno_remover = alunos_na_turma[escolha_aluno - 1]
                    nome_aluno, cpf_aluno = aluno_remover

                    turma_do_aluno = aluno_em_turma(book, (nome_aluno, cpf_aluno))
                    if turma_do_aluno:
                        aba_turma = book[turma_do_aluno]
                        for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                            nome = row[0].value
                            cpf = row[1].value
                            if (nome, cpf) == (nome_aluno, cpf_aluno):
                                aba_turma.delete_rows(row[0].row)
                                print(f"Aluno {nome_aluno} removido da turma {turma_do_aluno}.")
                                break
                    else:
                        print("Aluno não encontrado em nenhuma turma.")
                else:
                    print("Escolha de aluno inválida.")
            except ValueError:
                print("Escolha de aluno inválida.")
        else:
            print("Turma não encontrada ou não há alunos nessa turma.")

    elif escolha == '4':
        break
    else:
        print("\nOpção inválida. Escolha 1 para adicionar alunos, 2 para verificar alunos disponíveis, 3 para remover um aluno de uma turma, ou 4 para sair.")

print("\nAlterações realizadas com sucesso, encerrando o programa.", "\n")

# Salve as alterações no arquivo Excel
book.save(caminho_arquivo_excel)
