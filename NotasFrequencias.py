import openpyxl

# Função para registrar notas e frequência para um aluno
def registrar_notas_frequencia():
    rm_aluno = input("Digite o RM do aluno: ")
    codigo_professor = input("Digite o código do professor: ")
    
    wb = openpyxl.load_workbook('dados.xlsx')
    
    aluno_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name == "Alunos":
            aluno_sheet = wb[sheet_name]
            break
    
    if aluno_sheet is None:
        print("Planilha de alunos não encontrada.")
        return
    
    aluno_row = None
    for row in aluno_sheet.iter_rows(values_only=True):
        if row[0] == rm_aluno:
            aluno_row = row
            break
    
    if aluno_row is None:
        print("Aluno não encontrado.")
        return
    
    professor_sheet = wb["Prof"]
    
    if professor_sheet is None:
        print("Planilha de professores não encontrada.")
        return
    
    professor_encontrado = False
    for row in professor_sheet.iter_rows(values_only=True):
        if row[0] == codigo_professor:
            professor_encontrado = True
            break
    
    if not professor_encontrado:
        print("Professor não encontrado.")
        return
    
    num_notas = int(input("Digite a quantidade de notas a serem adicionadas: "))
    notas = []
    
    for i in range(num_notas):
        nota = float(input(f"Digite a nota {i + 1}: "))
        notas.append(nota)
    
    faltas = int(input("Digite o número de faltas do aluno: "))
    presencas = int(input("Digite o número de presenças do aluno: "))
    
    num_prazos = int(input("Digite a quantidade de prazos de entrega a serem adicionados: "))
    prazos = []
    
    for i in range(num_prazos):
        prazo = input(f"Digite o prazo {i + 1}: ")
        prazos.append(prazo)
    
    notas_str = ', '.join(map(str, notas))
    prazos_str = ', '.join(prazos)
    
    notas_sheet_name = "Notas"
    
    if notas_sheet_name not in wb.sheetnames:
        notas_sheet = wb.create_sheet(notas_sheet_name)
        notas_sheet['A1'] = "RM"
        notas_sheet['B1'] = "Notas"
        notas_sheet['C1'] = "Faltas"
        notas_sheet['D1'] = "Presenças"
        notas_sheet['E1'] = "Prazos de Entrega"
    else:
        notas_sheet = wb[notas_sheet_name]
    
    notas_sheet.append([rm_aluno, notas_str, faltas, presencas, prazos_str])
    
    wb.save('dados.xlsx')
    print(f"Notas, frequência e prazos de entrega registrados para o aluno com RM {rm_aluno} na aba 'Notas'.")

# Chamando a função para registrar notas e frequência
registrar_notas_frequencia()
