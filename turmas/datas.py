import openpyxl
from openpyxl.styles import Font, Alignment  # Importe as classes Font e Alignment
from datetime import datetime, timedelta
import os

# Construa o caminho completo para o arquivo Excel no diretório 'database'
diretorio_atual = os.path.dirname(os.path.abspath(__file__))
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

# Verificar se o arquivo Excel existe e criar um novo se não existir
if os.path.exists(caminho_arquivo_excel):
    planilha = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    planilha = openpyxl.Workbook()
    planilha.save(caminho_arquivo_excel)

def verificar_ciclo(n, x, y):
    if n > x or n < y:
        print("Data está fora do ciclo de curso, tente novamente")
    else:
        pass

def adicionar_data_e_ciclos(planilha, turma_destino):
    aba_turma = planilha[turma_destino]
    ciclos = []

    while True:
        try:
            data_inicio = datetime.strptime(input("\nDigite a data de início do curso (DD/MM/AAAA): "), "%d/%m/%Y")
            data_fim = datetime.strptime(input("Digite a data de término do curso (DD/MM/AAAA): "), "%d/%m/%Y")
            if data_fim < data_inicio:
                print("A data de término é anterior à data de início, tente novamente")
            else:
                break
        except ValueError:
            print("Formato de data inválido. Use o formato DD/MM/AAAA")

    aba_turma.cell(row=1, column=7).value = "Início do Curso"
    aba_turma.cell(row=1, column=8).value = "Fim do Curso"
    aba_turma.cell(row=2, column=7).value = data_inicio.strftime('%d/%m/%Y')
    aba_turma.cell(row=2, column=8).value = data_fim.strftime('%d/%m/%Y')

    aba_turma.cell(row=1, column=9).value = "Início do Ciclo"
    aba_turma.cell(row=1, column=10).value = "Término do Ciclo"
    aba_turma.cell(row=1, column=11).value = "Dias do Ciclo"

    qtd_ciclos = int(input("\nQuantos ciclos você deseja: "))

    while True:
        choice_cycle_type = input("\nEscolha o tipo do ciclo:\n\n1-Simétrico\n2-Definir cada ciclo\n\nEscolha uma das opções: ")
        if choice_cycle_type == "1" or choice_cycle_type == "2":
            break
        else:
            print("Opção inválida, tente novamente")

    if choice_cycle_type == "1":
        duracao_ciclo = (data_fim - data_inicio) / qtd_ciclos

        for i in range(qtd_ciclos - 1):
            ciclo_nome = f"Ciclo {i + 1}"
            ciclo_inicio = data_inicio + i * duracao_ciclo
            ciclo_fim = ciclo_inicio + duracao_ciclo - timedelta(days=1)
            ciclos.append((ciclo_nome, ciclo_inicio, ciclo_fim))

        ultimo_ciclo_nome = f"Ciclo {qtd_ciclos}"
        ultimo_ciclo_inicio = data_inicio + (qtd_ciclos - 1) * duracao_ciclo
        ultimo_ciclo_fim = data_fim
        ciclos.append((ultimo_ciclo_nome, ultimo_ciclo_inicio, ultimo_ciclo_fim))

    elif choice_cycle_type == "2":
        while True:
            try:
                duracao_ciclo = (data_fim - data_inicio)
                ciclo_datas = []  # Para armazenar as datas de início dos ciclos
                ciclo_datas_fim = []  # Para armazenar as datas de término dos ciclos

                for i in range(qtd_ciclos):
                    ciclo_nome = f"Ciclo {i + 1}"
                    while True:
                        try:
                            ciclo_inicio = datetime.strptime(input(f"\nDigite a data de início do {ciclo_nome} (DD/MM/AAAA): "), "%d/%m/%Y")
                            if ciclo_inicio < data_inicio or ciclo_inicio > data_fim:
                                print("Data está fora do ciclo de curso, tente novamente")
                            elif ciclo_inicio in ciclo_datas:
                                print("A data de início do ciclo já foi escolhida antes, tente novamente")
                            elif ciclo_inicio < max(ciclo_datas_fim, default=data_inicio):
                                print("A data de início do ciclo deve ser posterior à data de término do ciclo anterior.")
                            else:
                                break
                        except ValueError:
                            print("Formato de data inválido. Use o formato DD/MM/AAAA.")

                    while True:
                        try:
                            ciclo_fim = datetime.strptime(input(f"Digite a data de finalização do {ciclo_nome} (DD/MM/AAAA): "), "%d/%m/%Y")
                            if ciclo_fim < data_inicio or ciclo_fim > data_fim:
                                print("Data está fora do ciclo de curso, tente novamente")
                            elif ciclo_fim in ciclo_datas or ciclo_fim in ciclo_datas_fim:
                                print("A data de término do ciclo já foi escolhida antes, tente novamente")
                            elif ciclo_fim < ciclo_inicio:
                                print("A data de término do ciclo deve ser posterior à data de início do ciclo.")
                            else:
                                break
                        except ValueError:
                            print("Formato de data inválido. Use o formato DD/MM/AAAA.")

                    ciclos.append((ciclo_nome, ciclo_inicio, ciclo_fim))
                    ciclo_datas.append(ciclo_inicio)
                    ciclo_datas_fim.append(ciclo_fim)

                break
            except ValueError:
                print("Formato de data inválido. Use o formato DD/MM/AAAA.")

    for i, (ciclo_nome, ciclo_inicio, ciclo_fim) in enumerate(ciclos):
        aba_turma.cell(row=i + 2, column=9).value = ciclo_inicio.strftime('%d/%m/%Y')
        aba_turma.cell(row=i + 2, column=10).value = ciclo_fim.strftime('%d/%m/%Y')
        dias_ciclo = (ciclo_fim - ciclo_inicio).days + 1
        aba_turma.cell(row=i + 2, column=11).value = dias_ciclo
        print(f"\n{ciclo_nome} terá {dias_ciclo} dias.")

# Salvar a planilha no arquivo infodados.xlsx
    caminho_arquivo_excel = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'database', 'infodados.xlsx')
    planilha.save(caminho_arquivo_excel)
    print("\n----Data de início, fim do curso e ciclos adicionados com sucesso!!----")

# Abrir a planilha
caminho_arquivo_excel = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'database', 'infodados.xlsx')

while True:
    print("\nOpções:")
    print("\n1. Adicionar data de início do curso e criar ciclos")  
    print("2. Sair do programa", "\n")

    escolha = input("Escolha uma das opções: ")

    if escolha == '1':
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        if not abas_turmas:
            print("\nNão foram encontradas abas de turma na planilha")
        else:
            print("\nTurmas existentes:", "\n")
            for i, turma in enumerate(abas_turmas, start=1):
                print(f"{i}. {turma}")

            num_turma = int(input("\nEscolha o número da turma: "))
            if 1 <= num_turma <= len(abas_turmas):
                turma_desejada = abas_turmas[num_turma - 1]
                adicionar_data_e_ciclos(planilha, turma_desejada)
            else:
                print("\nNúmero de turma inválido.")

    elif escolha == '2':
        break
    else:
        print("Opção inválida. Escolha 1 para adicionar a data de início do curso e criar ciclos ou 2 para sair.")