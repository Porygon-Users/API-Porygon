import openpyxl
import os

# Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__ ))

# Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

# Abrir o arquivo Excel existente ou criar um novo
if os.path.exists(caminho_arquivo_excel):
    book = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    book = openpyxl.Workbook()

while True:
    try:
        # Carrega o arquivo
        book = openpyxl.load_workbook(caminho_arquivo_excel)
    except FileNotFoundError:
        print("O banco de dados não foi encontrado.")
        exit()

    def fechar_turma(book, turma_numero):
        turma_nome = f"Turma {turma_numero}"
        if turma_nome in book.sheetnames:
            sheet = book[turma_nome]
            novo_nome = f"{turma_nome} (fechada)"
            sheet.title = novo_nome
            book.save(caminho_arquivo_excel)  # Salva as alterações no arquivo Excel
            print(f"\n{turma_nome} foi fechada com sucesso.")
            return True
        else:
            print(f"\n\n\n\n\n\n\n\nDesculpe, a {turma_nome} já foi fechada ou não foi encontrada, tente novamente.")
            return False

    def listar_turmas_disponiveis(book):
        abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ') and not sheet.endswith('(fechada)')]
        print("\nTurmas disponíveis para fechamento:")
        for i, turma in enumerate(abas_turmas, start=1):
            print(f"{turma.replace('(fechada)', '')}")
        print()

    listar_turmas_disponiveis(book)
    turma_numero = int(input("Digite o número da turma que deseja fechar: "))

    if fechar_turma(book, turma_numero):
        while True:
            continuar = input("\nDeseja continuar fechando turmas? (S para continuar ou N para sair): ").strip().lower()
            if continuar == 's':
                break
            elif continuar == 'n':
                book.save(caminho_arquivo_excel)  # Salva as alterações antes de sair
                print("\n\nAlterações realizadas com sucesso, encerrando o programa.")
                exit()
            else:
                print("\n\nResposta inválida, por favor, tente novamente.")
