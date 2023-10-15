import openpyxl
from openpyxl.styles import Font, Alignment
import os
from openpyxl.utils import get_column_letter

# Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__))

# Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

# Abrir o arquivo Excel existente ou criar um novo
if os.path.exists(caminho_arquivo_excel):
    book = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    book = openpyxl.Workbook()

# Função para criar grupos em uma turma
def criar_grupos(planilha, turma_nome, num_alunos_por_grupo):
    # Carregar o arquivo da planilha
    wb = openpyxl.load_workbook(planilha)
    
    # Selecionar a aba da turma
    try:
        sheet = wb[turma_nome]
    except KeyError:
        print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
        return
    
    # Obter o índice da coluna que contém o cabeçalho "Grupos"
    coluna_grupos = None
    for cell in sheet[1]:  # Percorre as células da primeira linha
        if cell.value == "Grupos":
            coluna_grupos = cell.column_letter
            break
    
    if coluna_grupos is None:
        print("Não foi encontrada uma coluna com o cabeçalho 'Grupos'.")
        return
    
    # Obter a lista de alunos na turma
    alunos = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row if cell.value]
    
    # Exibir a quantidade de alunos na turma
    num_alunos = len(alunos)
    print(f"\nA turma '{turma_nome}' possui {num_alunos} alunos.")
    
    # Calcular o número total de grupos
    num_grupos = num_alunos // num_alunos_por_grupo
    alunos_restantes = num_alunos % num_alunos_por_grupo
    
    # Criar grupos com o número especificado de alunos
    grupos = []
    for i in range(num_grupos):
        grupo_alunos = alunos[i * num_alunos_por_grupo: (i + 1) * num_alunos_por_grupo]
        grupos.append(f'Grupo {i + 1} - {", ".join(grupo_alunos)}')
    
    # Se houver alunos restantes, criar o último grupo
    if alunos_restantes > 0:
        grupo_alunos = alunos[-alunos_restantes:]
        grupos.append(f'Grupo {num_grupos + 1} - {", ".join(grupo_alunos)}')
    
    # Adicionar os grupos à coluna "Grupos" na planilha
    for i, grupo in enumerate(grupos):
        sheet[f"{coluna_grupos}{i + 2}"] = grupo
    
    # Salvar a planilha
    wb.save(planilha)
    print(f"\nForam criados {len(grupos)} grupos com {num_alunos_por_grupo} alunos cada{' e um grupo final com ' + str(alunos_restantes) + ' alunos' if alunos_restantes > 0 else ''} na '{turma_nome}'.")

# Função para listar as turmas existentes
def listar_turmas(planilha):
    wb = openpyxl.load_workbook(planilha)
    turmas = [sheet for sheet in wb.sheetnames if sheet.startswith('Turma ')]
    if turmas:
        print("\nTurmas existentes:", "\n")
        for turma in turmas:
            print(turma)
    else:
        print("\nNão foram encontradas abas de turma na planilha.")

# Função para contar alunos em uma turma
def contar_alunos(planilha, turma_nome):
    # Carregar o arquivo da planilha
    wb = openpyxl.load_workbook(planilha)
    
    # Selecionar a aba da turma
    try:
        sheet = wb[turma_nome]
    except KeyError:
        print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
        return
    
    # Contar alunos na turma
    num_alunos = sum(1 for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) if row[0].value)
    return num_alunos  # Retornar a quantidade de alunos na turma

# Função principal do programa
def main():
    while True:
        print("\nOpções:")
        print("\n1. Criar grupos")
        print("2. Listar turmas existentes")
        print("3. Sair", "\n")
        
        escolha = input("Escolha uma opção: ")
        
        if escolha == '1':
            turma = input('\nDigite o nome da turma: ')
            alunos_na_turma = contar_alunos(caminho_arquivo_excel, turma)  # Mostrar a quantidade de alunos na turma
            if alunos_na_turma is not None:
                if alunos_na_turma > 0:
                    print(f"\nA turma '{turma}' possui {alunos_na_turma} alunos.")
                    alunos_por_grupo = int(input("\nDigite o número de alunos por grupo: "))
                    criar_grupos(caminho_arquivo_excel, turma, alunos_por_grupo)
        elif escolha == '2':
            listar_turmas(caminho_arquivo_excel)
        elif escolha == '3':
            print("\nSaindo do programa", "\n")
            break
        else:
            print("\nOpção inválida. Tente novamente.", "\n")

if __name__ == "__main__":
    main()
