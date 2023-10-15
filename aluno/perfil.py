import openpyxl
import os

# Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__))

# Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

# Verifique se o arquivo Excel existe
if os.path.exists(caminho_arquivo_excel):
    # Abrir o arquivo Excel existente
    book = openpyxl.load_workbook(caminho_arquivo_excel)

    # Solicitar login e senha do aluno
    login_aluno = input("Digite o login do aluno: ")
    senha_aluno = input("Digite a senha do aluno: ")

    aluno_encontrado = None

    # Primeiro, procurar o aluno na aba de cadastro
    aba_cadastro = book['Cadastro']
    for row in aba_cadastro.iter_rows(min_row=2, values_only=True):
        if row[3] == login_aluno and row[4] == senha_aluno:  # Usar índices 3 e 4 para D e E
            aluno_encontrado = {
                "nome": row[0],
                "cpf": row[1]
            }
            break

    if aluno_encontrado:
        print("Aluno encontrado na aba de cadastro.")

        # Agora, procure o aluno nas outras abas
        for sheet_name in book.sheetnames:
            if sheet_name != 'Cadastro':  # Ignorar a aba de cadastro
                sheet = book[sheet_name]
                aluno_presente = False
                # Procurar pelo aluno com o nome e CPF fornecidos nas outras abas
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] == aluno_encontrado["nome"] and row[1] == aluno_encontrado["cpf"]:
                        aluno_presente = True
                        print(f"Aluno encontrado na planilha '{sheet_name}'.")
                        # Exibir outras informações da planilha e do aluno conforme necessário
                        break

                if not aluno_presente:
                    print(f"Aluno não encontrado na planilha '{sheet_name}'.")

    else:
        print("Aluno não encontrado na aba de cadastro.")

    # Fechar o arquivo Excel
    book.close()

else:
    print(f"O arquivo {caminho_arquivo_excel} não existe.")
