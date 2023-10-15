import openpyxl
import os
import random

# Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__))

# Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

# Abrir o arquivo Excel existente ou criar um novo
if os.path.exists(caminho_arquivo_excel):
    arquivo_excel = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    arquivo_excel = openpyxl.Workbook()

# Gera uma senha aleatória com letras maiúsculas, minúsculas e números, contendo no máximo 7 caracteres
def gerar_senha():
    caracteres = "1234567890ABCDEF"
    senha = ''.join(random.choice(caracteres) for i in range(7))
    return senha

def gerar_senha1():
    caracteres = "abcdefgABCDEFG01234567"
    senha = ''.join(random.choice(caracteres) for i in range(7))
    return senha

# Validação de CPF como números com 11 caracteres
def validar_cpf(cpf):
    if cpf.isdigit() and len(cpf) == 11:
        return True
    else:
        return False

# Formata o CPF com os "-" nos dois últimos dígitos
def formatar_cpf(cpf):
    return f"{cpf[:9]}-{cpf[9:]}"

# Obtenha a planilha ou crie uma nova se não existir
if "Cadastro" in arquivo_excel.sheetnames:
    planilha = arquivo_excel["Cadastro"]
else:
    planilha = arquivo_excel.active
    planilha.title = "Cadastro"
    planilha['A1'] = "Função"
    planilha['B1'] = "Nome"
    planilha['C1'] = "Email"
    planilha['D1'] = "CPF"

while True:
    print("\nBem vindo(a) ao sistema de cadastro da PBLTeX!", "\n")
    funcao = input("Gostaria de se cadastrar como 'Aluno' ou 'Professor'? Para sair, digite 'sair': ").lower()
    if funcao == 'sair':
        print("\nEspero te ver novamente, até logo! ", "\n")
        break
    if funcao not in ['aluno', 'professor']:
        print("Opção inválida. Tente novamente.")
        continue

    nome = input("Digite o nome: ").lower()
    telefone = input("Digite seu telefone com o DDD (Ex: 12912345678): ")

    # Obtém o primeiro nome e as iniciais do último nome
    partes_nome = nome.split()
    primeiro_nome = partes_nome[0]
    iniciais_ultimo_nome = "".join([partes_nome[-1][0]])

    # Construa o login e o email usando o primeiro nome e as iniciais do último nome
    login = primeiro_nome + iniciais_ultimo_nome + str(random.randint(1, 10))
    email = primeiro_nome + iniciais_ultimo_nome + "@pbltex.com"

    # Validação de CPF como números com 11 caracteres
    cpf_valido = False  # Inicializa como falso
    while not cpf_valido:  # Continua solicitando CPF até que seja válido
        cpf = input("Digite o CPF: ")
        if validar_cpf(cpf):
            cpf = formatar_cpf(cpf)  # Formata o CPF com os "-"
            cpf_valido = True  # Define como verdadeiro para sair do loop
        else:
            print("\nCPF inválido, tente novamente.")

    if funcao == 'Aluno':
        senha = gerar_senha()
    else:
        senha = gerar_senha1()  # Senha para professores é diferente da do aluno, mas com o mesmo princípio.
    # Encontre a próxima linha vazia na planilha
    proxima_linha = planilha.max_row + 1

    # Preencha os dados nas colunas apropriadas
    planilha.cell(row=proxima_linha, column=4, value=funcao)
    planilha.cell(row=proxima_linha, column=1, value=nome)
    planilha.cell(row=proxima_linha, column=3, value=email)
    planilha.cell(row=proxima_linha, column=2, value=cpf)

    # Salve o login e a senha nas colunas apropriadas
    planilha.cell(row=proxima_linha, column=5, value=login)  # Coluna 5 para o login
    planilha.cell(row=proxima_linha, column=6, value=senha)  # Coluna 6 para a senha

    print("\nBem vindo(a) a PBLTeX.", "\n")
    print(f"E-mail Institucional: {email}")
    print(f"Login: {login}")
    print(f"Senha: {senha}", "\n")
    print("Cadastro feito com sucesso! Foi enviado no número de telefone informado, seu usuário e senha para efetuar o login.", "\n")
    break  # Encerra o programa após o cadastro

# Salve as alterações no arquivo Excel
arquivo_excel.save(caminho_arquivo_excel)
arquivo_excel.close()
